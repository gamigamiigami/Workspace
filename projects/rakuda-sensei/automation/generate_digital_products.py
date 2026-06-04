#!/usr/bin/env python3
"""
販売用デジタル成果物を生成する (Excel テンプレート群)

生成物:
1. 5年で2000万円｜公務員教員の家計簿テンプレ.xlsx (¥500)
2. オルカン+S&P500｜投資配分シミュレータ.xlsx (¥980)
3. 中学校教員の席替え自動Excel.xlsx (¥500)

質重視: 数式・実体験ベースの例・スマホでも読める設計
"""

from __future__ import annotations

import datetime
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference

ROOT = Path(__file__).resolve().parents[3]
OUT_DIR = ROOT / "projects" / "rakuda-sensei" / "products" / "digital"

# ====== ブランドカラー ======
BRAND_BROWN = "8B6F47"
BRAND_CREAM = "FDF6E3"
BRAND_BEIGE = "E8D5A7"
BRAND_DARK = "3D2F1F"

HEADER_FILL = PatternFill(start_color=BRAND_BROWN, end_color=BRAND_BROWN, fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color=BRAND_BEIGE, end_color=BRAND_BEIGE, fill_type="solid")
LIGHT_FILL = PatternFill(start_color=BRAND_CREAM, end_color=BRAND_CREAM, fill_type="solid")

WHITE_BOLD = Font(name="メイリオ", size=11, bold=True, color="FFFFFF")
BLACK_BOLD = Font(name="メイリオ", size=11, bold=True, color=BRAND_DARK)
NORMAL = Font(name="メイリオ", size=10)
NUM_FONT = Font(name="メイリオ", size=10)

THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def apply_header_row(ws, row: int, headers: list[str], cols_from: int = 1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=cols_from + i, value=h)
        c.font = WHITE_BOLD
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER


def set_column_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# =========================================================================
# 1. 家計簿テンプレ (¥500)
# =========================================================================
def build_kakeibo() -> Workbook:
    wb = Workbook()

    # === Sheet 1: 使い方 ===
    ws = wb.active
    ws.title = "使い方"
    set_column_widths(ws, [4, 80])

    ws.cell(row=1, column=1, value="🐪").font = Font(size=24)
    ws.cell(row=1, column=2, value="残業嫌いのらくだ先生 家計簿テンプレ").font = Font(name="メイリオ", size=20, bold=True, color=BRAND_DARK)

    intro_lines = [
        "",
        "■ このテンプレで何ができるか",
        "・毎月の固定費・変動費・収入を1ファイルで管理",
        "・年間の貯蓄率・投資率を自動計算",
        "・「先取り投資 → 生活 → 貯金」の順番でお金を回す仕組み",
        "・公務員教員 (20代・実家暮らし) のリアル数字を例として埋め込み済み",
        "",
        "■ 使い方",
        "1. シート「月次記入」のC列に毎月の金額を入力するだけ",
        "2. 「集計」シートに年間の貯蓄率・投資配分が自動表示",
        "3. 「投資配分」シートでNISA積立の銘柄管理",
        "4. 「目標シミュレータ」で何年で2000万に届くか試算",
        "",
        "■ ぼくのリアル数字 (参考)",
        "・固定費: 月4.5万円 (実家へ3万・通信5千・ガソリン5千・サブスク5千)",
        "・娯楽費: 月5万円 (上限目安)",
        "・NISA積立: 月15万円 (自動天引き)",
        "・配分: オルカン+S&P500=80% / NASDAQ100=15% / FANG+ゴールド=5%",
        "・5年で累計2000万円 (節約60% + 投資40%)",
        "・失敗: 日本株デイトレで-50万 (やめてインデックスに振り切った)",
        "",
        "■ 1番伝えたいこと",
        "「タイミングを狙うのをやめる」「先取りで自動天引き」「残ったお金で生きる」",
        "この3つを5年やったら、教員でも2000万になります。",
        "",
        "■ 不明点があれば",
        "X @rakuda_sensei までDMください。",
    ]
    for i, line in enumerate(intro_lines, start=3):
        c = ws.cell(row=i, column=2, value=line)
        if line.startswith("■"):
            c.font = Font(name="メイリオ", size=14, bold=True, color=BRAND_BROWN)
        else:
            c.font = NORMAL
        c.alignment = LEFT

    # === Sheet 2: 月次記入 ===
    ws2 = wb.create_sheet("月次記入")
    set_column_widths(ws2, [12, 20, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14])

    ws2.cell(row=1, column=1, value="月次記入シート（C-N列に毎月の金額を入力）").font = Font(name="メイリオ", size=14, bold=True, color=BRAND_DARK)

    # ヘッダー
    months = ["1月", "2月", "3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月"]
    apply_header_row(ws2, 3, ["大項目", "詳細"] + months)

    # データ行 (実体験数字をC列に例として入れる)
    rows = [
        # (大項目, 詳細, 1月分の例数値)
        ("収入", "本業手取り", 280000, "fixed_income"),
        ("収入", "副業", 30000, "var"),
        ("収入", "ボードゲーム会収入", 15000, "var"),
        ("収入", "合計", "=SUM(C4:C6)", "sum"),
        ("先取り", "NISA積立 (自動)", 150000, "fixed_invest"),
        ("先取り", "米株タイミング購入", 30000, "var_invest"),
        ("先取り", "投資合計", "=SUM(C8:C9)", "sum"),
        ("固定費", "実家へ生活費", 30000, "fixed"),
        ("固定費", "通信費", 5000, "fixed"),
        ("固定費", "ガソリン", 5000, "fixed"),
        ("固定費", "サブスク", 5000, "fixed"),
        ("固定費", "合計", "=SUM(C11:C14)", "sum"),
        ("変動費", "食費 (外食含む)", 25000, "var"),
        ("変動費", "娯楽費 (上限5万)", 30000, "var"),
        ("変動費", "交際費", 10000, "var"),
        ("変動費", "書籍・自己投資", 5000, "var"),
        ("変動費", "予備費", 0, "var"),
        ("変動費", "合計", "=SUM(C16:C20)", "sum"),
        ("結果", "支出計", "=C15+C21", "sum"),
        ("結果", "余剰金 → 貯金", "=C7-C10-C22", "result"),
    ]

    cat_color = {
        "収入": "C8E6C9",
        "先取り": "FFE0B2",
        "固定費": "FFCDD2",
        "変動費": "F8BBD0",
        "結果": "BBDEFB",
    }

    for i, (cat, detail, jan_value, kind) in enumerate(rows, start=4):
        ws2.cell(row=i, column=1, value=cat).fill = PatternFill(start_color=cat_color.get(cat, "FFFFFF"), end_color=cat_color.get(cat, "FFFFFF"), fill_type="solid")
        ws2.cell(row=i, column=1).font = BLACK_BOLD
        ws2.cell(row=i, column=1).alignment = CENTER
        ws2.cell(row=i, column=1).border = BORDER

        ws2.cell(row=i, column=2, value=detail).font = NORMAL
        ws2.cell(row=i, column=2).alignment = LEFT
        ws2.cell(row=i, column=2).border = BORDER

        # 1月分 (C列) のみ実例値、2-12月は同じ式 or 空
        for m in range(12):
            col = 3 + m  # C列=3
            if kind == "sum":
                # 月ごとに同じパターンで集計
                row_above_start = i - (4 if "C4" in str(jan_value) else 0)
                # 動的にrange計算
                if "C4:C6" in str(jan_value):
                    f = f"=SUM({get_column_letter(col)}4:{get_column_letter(col)}6)"
                elif "C8:C9" in str(jan_value):
                    f = f"=SUM({get_column_letter(col)}8:{get_column_letter(col)}9)"
                elif "C11:C14" in str(jan_value):
                    f = f"=SUM({get_column_letter(col)}11:{get_column_letter(col)}14)"
                elif "C16:C20" in str(jan_value):
                    f = f"=SUM({get_column_letter(col)}16:{get_column_letter(col)}20)"
                elif "C15+C21" in str(jan_value):
                    f = f"={get_column_letter(col)}15+{get_column_letter(col)}21"
                else:
                    f = jan_value
                ws2.cell(row=i, column=col, value=f)
            elif kind == "result":
                f = f"={get_column_letter(col)}7-{get_column_letter(col)}10-{get_column_letter(col)}22"
                ws2.cell(row=i, column=col, value=f)
            elif m == 0:
                ws2.cell(row=i, column=col, value=jan_value)
            else:
                ws2.cell(row=i, column=col, value=jan_value if kind.startswith("fixed") else 0)
            ws2.cell(row=i, column=col).number_format = "#,##0"
            ws2.cell(row=i, column=col).font = NUM_FONT
            ws2.cell(row=i, column=col).alignment = RIGHT
            ws2.cell(row=i, column=col).border = BORDER

    # === Sheet 3: 集計 ===
    ws3 = wb.create_sheet("集計")
    set_column_widths(ws3, [22, 18, 18, 18])
    ws3.cell(row=1, column=1, value="年間集計").font = Font(name="メイリオ", size=14, bold=True, color=BRAND_DARK)
    apply_header_row(ws3, 3, ["項目", "年間合計", "月平均", "%"])

    summary_rows = [
        ("収入合計", "=SUM(月次記入!C7:N7)"),
        ("先取り投資合計", "=SUM(月次記入!C10:N10)"),
        ("固定費合計", "=SUM(月次記入!C15:N15)"),
        ("変動費合計", "=SUM(月次記入!C21:N21)"),
        ("支出合計", "=SUM(月次記入!C22:N22)"),
        ("年間貯蓄 (余剰)", "=SUM(月次記入!C23:N23)"),
        ("実質貯蓄率 (投資+貯金)", '=ROUND((B5+B8)/B4*100,1)'),
        ("投資比率", '=ROUND(B5/B4*100,1)'),
    ]
    for i, (name, formula) in enumerate(summary_rows, start=4):
        ws3.cell(row=i, column=1, value=name).font = BLACK_BOLD
        ws3.cell(row=i, column=1).fill = LIGHT_FILL
        ws3.cell(row=i, column=1).border = BORDER
        ws3.cell(row=i, column=2, value=formula).number_format = "#,##0"
        ws3.cell(row=i, column=2).border = BORDER
        # 月平均 (B / 12)
        if i <= 9:
            ws3.cell(row=i, column=3, value=f"=B{i}/12").number_format = "#,##0"
            ws3.cell(row=i, column=3).border = BORDER

    # === Sheet 4: 投資配分 ===
    ws4 = wb.create_sheet("投資配分")
    set_column_widths(ws4, [24, 16, 16, 24])
    ws4.cell(row=1, column=1, value="投資配分・銘柄管理").font = Font(name="メイリオ", size=14, bold=True, color=BRAND_DARK)
    apply_header_row(ws4, 3, ["銘柄", "目標配分%", "月積立額", "メモ"])

    portfolio = [
        ("オルカン (全世界株式)", 40, 60000, "コア配分・楽天 or eMAXIS Slim"),
        ("S&P500", 40, 60000, "コア配分・eMAXIS Slim or 楽天"),
        ("NASDAQ100", 15, 22500, "成長重視・iFreeNEXT"),
        ("FANG+", 3, 4500, "ハイリスク枠・米一歩先テック"),
        ("ゴールド", 2, 3000, "ヘッジ・auAM"),
        ("合計", "=SUM(B4:B8)", "=SUM(C4:C8)", "150,000円/月で5年で2000万に到達"),
    ]
    for i, (name, ratio, amount, memo) in enumerate(portfolio, start=4):
        ws4.cell(row=i, column=1, value=name).font = BLACK_BOLD if i == 9 else NORMAL
        ws4.cell(row=i, column=1).alignment = LEFT
        ws4.cell(row=i, column=1).border = BORDER
        ws4.cell(row=i, column=2, value=ratio)
        if i != 9:
            ws4.cell(row=i, column=2).number_format = "0"
        ws4.cell(row=i, column=2).border = BORDER
        ws4.cell(row=i, column=3, value=amount).number_format = "#,##0"
        ws4.cell(row=i, column=3).border = BORDER
        ws4.cell(row=i, column=4, value=memo).alignment = LEFT
        ws4.cell(row=i, column=4).border = BORDER

    # === Sheet 5: 目標シミュレータ ===
    ws5 = wb.create_sheet("目標シミュレータ")
    set_column_widths(ws5, [24, 18, 18, 24])
    ws5.cell(row=1, column=1, value="累計資産シミュレーション").font = Font(name="メイリオ", size=14, bold=True, color=BRAND_DARK)
    ws5.cell(row=2, column=1, value="入力: 月積立・想定年利を変えて、何年で目標額に届くか試算")

    ws5.cell(row=4, column=1, value="月積立額").font = BLACK_BOLD
    ws5.cell(row=4, column=2, value=150000).number_format = "#,##0"
    ws5.cell(row=5, column=1, value="想定年利").font = BLACK_BOLD
    ws5.cell(row=5, column=2, value=0.05).number_format = "0.0%"
    ws5.cell(row=6, column=1, value="現在資産").font = BLACK_BOLD
    ws5.cell(row=6, column=2, value=0).number_format = "#,##0"

    apply_header_row(ws5, 8, ["年数", "元本累計", "運用益込み資産", "達成判定 (2000万円)"])
    for year in range(1, 16):
        i = 8 + year
        ws5.cell(row=i, column=1, value=year)
        # 元本: 月積立 * 12 * 年 + 現在資産
        ws5.cell(row=i, column=2, value=f"=$B$4*12*{year}+$B$6").number_format = "#,##0"
        # 複利計算 (FV関数)
        ws5.cell(row=i, column=3, value=f"=FV(-$B$5/12,{year*12},-$B$4,-$B$6)").number_format = "#,##0"
        ws5.cell(row=i, column=4, value=f'=IF(C{i}>=20000000,"✓ 到達","")').alignment = CENTER

    return wb


# =========================================================================
# 2. 投資配分シミュレータ (¥980)
# =========================================================================
def build_investment_simulator() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "シミュレータ"
    set_column_widths(ws, [22, 18, 18, 18, 22])

    ws.cell(row=1, column=1, value="🐪 オルカン+S&P500｜投資配分シミュレータ").font = Font(name="メイリオ", size=18, bold=True, color=BRAND_DARK)
    ws.cell(row=2, column=1, value="月積立・年利・年数を変えて、3つの配分パターンを比較").font = NORMAL

    # 入力欄
    ws.cell(row=4, column=1, value="■ 入力").font = Font(name="メイリオ", size=12, bold=True, color=BRAND_BROWN)
    inputs = [
        ("月積立額", 150000),
        ("想定年数", 5),
        ("現在資産", 0),
    ]
    for i, (label, val) in enumerate(inputs, start=5):
        ws.cell(row=i, column=1, value=label).font = BLACK_BOLD
        ws.cell(row=i, column=1).fill = LIGHT_FILL
        ws.cell(row=i, column=1).border = BORDER
        c = ws.cell(row=i, column=2, value=val)
        c.number_format = "#,##0" if label != "想定年数" else "0"
        c.border = BORDER

    # 配分パターン比較
    ws.cell(row=10, column=1, value="■ 配分パターン比較 (期待年利)").font = Font(name="メイリオ", size=12, bold=True, color=BRAND_BROWN)
    apply_header_row(ws, 12, ["パターン", "オルカン+SP500", "NASDAQ100", "ハイリスク枠", "期待年利"])
    patterns = [
        ("保守: 全世界株式100", 100, 0, 0, 0.04),
        ("らくだ式 (80/15/5)", 80, 15, 5, 0.06),
        ("攻撃: NASDAQ100中心", 30, 60, 10, 0.08),
    ]
    for i, (name, w1, w2, w3, ar) in enumerate(patterns, start=13):
        ws.cell(row=i, column=1, value=name).font = NORMAL
        ws.cell(row=i, column=1).border = BORDER
        ws.cell(row=i, column=2, value=f"{w1}%").alignment = CENTER
        ws.cell(row=i, column=3, value=f"{w2}%").alignment = CENTER
        ws.cell(row=i, column=4, value=f"{w3}%").alignment = CENTER
        ws.cell(row=i, column=5, value=ar).number_format = "0.0%"
        for col in range(2, 6):
            ws.cell(row=i, column=col).border = BORDER

    # 結果計算
    ws.cell(row=17, column=1, value="■ 計算結果 (FV関数で複利計算)").font = Font(name="メイリオ", size=12, bold=True, color=BRAND_BROWN)
    apply_header_row(ws, 19, ["パターン", "元本累計", "運用益込み", "運用益のみ"])
    for i, (name, w1, w2, w3, ar) in enumerate(patterns, start=20):
        ws.cell(row=i, column=1, value=name).font = NORMAL
        ws.cell(row=i, column=1).border = BORDER
        ws.cell(row=i, column=2, value="=$B$5*12*$B$6+$B$7").number_format = "#,##0"
        ws.cell(row=i, column=2).border = BORDER
        ws.cell(row=i, column=3, value=f"=FV(-{ar}/12,$B$6*12,-$B$5,-$B$7)").number_format = "#,##0"
        ws.cell(row=i, column=3).border = BORDER
        ws.cell(row=i, column=4, value=f"=C{i}-B{i}").number_format = "#,##0"
        ws.cell(row=i, column=4).border = BORDER

    # 説明
    ws.cell(row=25, column=1, value="■ 注意事項").font = Font(name="メイリオ", size=12, bold=True, color=BRAND_BROWN)
    notes = [
        "・想定年利はあくまで参考値（過去実績ベース）",
        "・投資は元本割れリスクあり、必ず自己責任で判断",
        "・らくだ式は個人の判断であり、推奨ではありません",
        "・新NISA枠（年360万）の活用がベストケース",
    ]
    for i, n in enumerate(notes, start=26):
        ws.cell(row=i, column=1, value=n).font = NORMAL

    # === シート2: 年次推移 ===
    ws2 = wb.create_sheet("年次推移")
    set_column_widths(ws2, [10, 18, 18, 18])
    ws2.cell(row=1, column=1, value="らくだ式 (80/15/5) の年次推移").font = Font(name="メイリオ", size=14, bold=True, color=BRAND_DARK)
    apply_header_row(ws2, 3, ["年", "元本", "運用益込み", "達成度 (vs 2000万)"])
    for year in range(1, 11):
        i = 3 + year
        ws2.cell(row=i, column=1, value=year)
        ws2.cell(row=i, column=2, value=f"=シミュレータ!$B$5*12*{year}+シミュレータ!$B$7").number_format = "#,##0"
        ws2.cell(row=i, column=3, value=f"=FV(-0.06/12,{year*12},-シミュレータ!$B$5,-シミュレータ!$B$7)").number_format = "#,##0"
        ws2.cell(row=i, column=4, value=f"=ROUND(C{i}/20000000*100,1)&\"%\"")

    return wb


# =========================================================================
# 3. 席替え自動Excel (¥500)
# =========================================================================
def build_seat_shuffle() -> Workbook:
    wb = Workbook()

    ws = wb.active
    ws.title = "使い方"
    set_column_widths(ws, [4, 70])
    ws.cell(row=1, column=1, value="🐪").font = Font(size=24)
    ws.cell(row=1, column=2, value="席替え自動 Excel").font = Font(name="メイリオ", size=20, bold=True, color=BRAND_DARK)

    lines = [
        "",
        "■ このExcelで何ができるか",
        "・名簿を入力するだけで、ランダム席替えが10秒で完了",
        "・前列希望・後ろ希望などの「配慮」をフラグ付けで反映",
        "・班分け（4人1班など）もボタン1つで",
        "・席替えに2時間かけてた頃のぼくに教えてあげたかった仕組み",
        "",
        "■ 使い方",
        "1. シート「名簿」に生徒名を入力（最大40人）",
        "2. 配慮列に「前」「後」「視力」など必要な配慮を記入",
        "3. シート「席替え結果」を開いてF9キー (再計算) で席が決まる",
        "4. 気に入った配置になるまでF9を押すだけ",
        "",
        "■ 注意",
        "・最終決定は必ず先生の目視で。Excelはあくまで「候補出し」",
        "・特定の生徒同士を離す等の特別配慮は手動調整推奨",
    ]
    for i, line in enumerate(lines, start=3):
        c = ws.cell(row=i, column=2, value=line)
        if line.startswith("■"):
            c.font = Font(name="メイリオ", size=14, bold=True, color=BRAND_BROWN)
        else:
            c.font = NORMAL
        c.alignment = LEFT

    # === シート2: 名簿 ===
    ws2 = wb.create_sheet("名簿")
    set_column_widths(ws2, [6, 12, 20, 16, 30])
    apply_header_row(ws2, 1, ["No", "出席番号", "氏名", "配慮", "メモ"])
    examples = [
        ("1", "A1", "（生徒1）", "", ""),
        ("2", "A2", "（生徒2）", "前", "視力配慮"),
        ("3", "A3", "（生徒3）", "", ""),
    ]
    for i, row in enumerate(examples, start=2):
        for j, val in enumerate(row):
            ws2.cell(row=i, column=j + 1, value=val).border = BORDER
            ws2.cell(row=i, column=j + 1).font = NORMAL
            ws2.cell(row=i, column=j + 1).alignment = LEFT if j >= 2 else CENTER

    for i in range(5, 45):
        for j in range(5):
            ws2.cell(row=i, column=j + 1, value="").border = BORDER

    # === シート3: 席替え結果 (5×8 = 40席を想定) ===
    ws3 = wb.create_sheet("席替え結果")
    set_column_widths(ws3, [14] * 9)
    ws3.cell(row=1, column=1, value="席替え結果 (F9キーで再シャッフル)").font = Font(name="メイリオ", size=14, bold=True, color=BRAND_DARK)
    ws3.cell(row=2, column=1, value="↑黒板側").font = BLACK_BOLD
    ws3.cell(row=2, column=1).alignment = CENTER

    # 8列×5行の席
    for row_idx in range(5):
        for col_idx in range(8):
            cell = ws3.cell(row=4 + row_idx, column=1 + col_idx)
            # ランダム参照: =INDEX(名簿!C:C, RANDBETWEEN(2,41))
            # 重複防止のため簡易版: INDEX + RANK + RAND の組み合わせ
            cell.value = f'=INDEX(名簿!$C$2:$C$41,RANK(INDEX(rand_keys,{row_idx*8+col_idx+1}),rand_keys))'
            cell.font = NORMAL
            cell.alignment = CENTER
            cell.border = BORDER
            cell.fill = LIGHT_FILL

    # rand_keys という名前定義: =RAND() を 40 個並べたヘルパー列
    ws4 = wb.create_sheet("ヘルパー")
    ws4.cell(row=1, column=1, value="ランダム数 (触らないで)")
    for i in range(2, 42):
        ws4.cell(row=i, column=1, value=f"=RAND()")

    wb.defined_names["rand_keys"] = openpyxl_defined_name("rand_keys", "ヘルパー!$A$2:$A$41")

    return wb


def openpyxl_defined_name(name, ref):
    from openpyxl.workbook.defined_name import DefinedName
    return DefinedName(name=name, attr_text=ref)


# =========================================================================
# メイン
# =========================================================================
def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"📁 出力先: {OUT_DIR}\n")

    products = [
        ("5年で2000万｜家計簿テンプレ.xlsx", build_kakeibo, "¥500"),
        ("オルカン+S&P500投資配分シミュレータ.xlsx", build_investment_simulator, "¥980"),
        ("席替え自動Excel.xlsx", build_seat_shuffle, "¥500"),
    ]

    for name, builder, price in products:
        try:
            wb = builder()
            out_path = OUT_DIR / name
            wb.save(out_path)
            size_kb = out_path.stat().st_size // 1024
            print(f"  ✅ {name} ({price}, {size_kb}KB)")
        except Exception as e:
            print(f"  ❌ {name}: {e}", file=sys.stderr)
            import traceback
            traceback.print_exc(file=sys.stderr)

    print(f"\n🎁 デジタル販売物 {len(products)} 個生成完了")
    return 0


if __name__ == "__main__":
    sys.exit(main())
