#!/usr/bin/env python3
"""
サイドFIRE計画シート v3 (入力集約 + 結論一目瞭然)

設計原則:
- 入力セル: 黄色塗りつぶし + 太枠 (それ以外は塗りつぶし無し)
- 入力は1シートに全集約
- 「結論」シートで「サイドFIRE実現可否・不足額・達成年齢」が即わかる
- 補助シートは年次推移とFIRE試算のみ

サイドFIRE 4%ルール:
  必要資産 = 年生活費(夫婦) × 25 -- フルFIRE
  サイドFIRE = (年生活費 - 副業年収) × 25
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.defined_name import DefinedName

ROOT = Path(__file__).resolve().parents[3]
OUT_DIR = ROOT / "projects" / "rakuda-sensei" / "products" / "digital"

# ====== カラーと書式 ======
BRAND_BROWN = "8B6F47"
BRAND_BEIGE = "E8D5A7"
BRAND_DARK = "3D2F1F"

INPUT_YELLOW = "FFF59D"  # 鮮やかな黄色 (入力セル明確化)
OUTPUT_GREEN = "A5D6A7"  # 結論セル
WARNING_RED = "EF9A9A"   # 警告セル
LABEL_WHITE = "FFFFFF"

def F(size=11, bold=False, color=BRAND_DARK, italic=False):
    return Font(name="メイリオ", size=size, bold=bold, color=color, italic=italic)

TITLE = F(size=20, bold=True)
H1 = F(size=14, bold=True, color="FFFFFF")
H2 = F(size=12, bold=True, color=BRAND_BROWN)
LABEL = F(size=11, bold=True)
NORMAL = F(size=11)
SMALL = F(size=10, color="666666")
BIG_RESULT = F(size=18, bold=True, color=BRAND_DARK)
TINY_NOTE = F(size=9, italic=True, color="999999")

FILL_HEADER = PatternFill(start_color=BRAND_BROWN, end_color=BRAND_BROWN, fill_type="solid")
FILL_SECTION = PatternFill(start_color=BRAND_BEIGE, end_color=BRAND_BEIGE, fill_type="solid")
FILL_INPUT = PatternFill(start_color=INPUT_YELLOW, end_color=INPUT_YELLOW, fill_type="solid")
FILL_OUTPUT = PatternFill(start_color=OUTPUT_GREEN, end_color=OUTPUT_GREEN, fill_type="solid")
FILL_WARN = PatternFill(start_color=WARNING_RED, end_color=WARNING_RED, fill_type="solid")
FILL_NONE = PatternFill(fill_type=None)

T = Side(border_style="thin", color="CCCCCC")
M = Side(border_style="medium", color=BRAND_BROWN)
BORDER_THIN = Border(left=T, right=T, top=T, bottom=T)
BORDER_INPUT = Border(left=M, right=M, top=M, bottom=M)
BORDER_OUTPUT = Border(left=M, right=M, top=M, bottom=M)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def title_row(ws, title: str, sub: str = ""):
    ws.cell(row=1, column=1, value="🐪").font = F(size=24)
    c = ws.cell(row=1, column=2, value=title)
    c.font = TITLE
    ws.row_dimensions[1].height = 34
    if sub:
        s = ws.cell(row=2, column=2, value=sub)
        s.font = SMALL


def section(ws, row: int, text: str, span: int = 5):
    for c in range(1, span + 1):
        ws.cell(row=row, column=c).fill = FILL_SECTION
        ws.cell(row=row, column=c).border = BORDER_THIN
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = H2
    cell.alignment = LEFT
    ws.row_dimensions[row].height = 24


def input_field(ws, row: int, label: str, value, fmt: str = "#,##0", note: str = ""):
    """入力フィールド: ラベル(B列) + 入力値(C列) + 注記(D列以降)"""
    lc = ws.cell(row=row, column=2, value=label)
    lc.font = LABEL
    lc.alignment = LEFT
    lc.border = BORDER_THIN

    ic = ws.cell(row=row, column=3, value=value)
    ic.fill = FILL_INPUT  # 黄色塗りつぶし
    ic.border = BORDER_INPUT
    ic.font = F(size=12, bold=True)
    ic.alignment = RIGHT
    ic.number_format = fmt

    if note:
        nc = ws.cell(row=row, column=4, value=note)
        nc.font = TINY_NOTE
        nc.alignment = LEFT


def calc_field(ws, row: int, label: str, formula, fmt: str = "#,##0", note: str = ""):
    """計算結果: ラベル + 結果 (色なし)"""
    lc = ws.cell(row=row, column=2, value=label)
    lc.font = NORMAL
    lc.alignment = LEFT
    lc.border = BORDER_THIN

    cc = ws.cell(row=row, column=3, value=formula)
    cc.border = BORDER_THIN
    cc.font = NORMAL
    cc.alignment = RIGHT
    cc.number_format = fmt

    if note:
        nc = ws.cell(row=row, column=4, value=note)
        nc.font = SMALL
        nc.alignment = LEFT


def result_field(ws, row: int, label: str, formula, fmt: str = "#,##0", note: str = ""):
    """重要アウトプット: 緑塗りつぶし"""
    lc = ws.cell(row=row, column=2, value=label)
    lc.font = LABEL
    lc.alignment = LEFT
    lc.border = BORDER_OUTPUT

    rc = ws.cell(row=row, column=3, value=formula)
    rc.fill = FILL_OUTPUT
    rc.border = BORDER_OUTPUT
    rc.font = BIG_RESULT
    rc.alignment = RIGHT
    rc.number_format = fmt
    ws.row_dimensions[row].height = 28

    if note:
        nc = ws.cell(row=row, column=4, value=note)
        nc.font = SMALL
        nc.alignment = LEFT


# =========================================================================
def build():
    wb = Workbook()
    wb.remove(wb.active)

    # ============ 使い方 ============
    ws = wb.create_sheet("使い方", 0)
    ws.sheet_properties.tabColor = BRAND_BROWN
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 95
    title_row(ws, "夫婦で実践｜サイドFIRE計画シート v3",
              "1シートに入力 → 結論シートで実現可否が即わかる")
    lines = [
        "",
        "■ シート構成（読む順番）",
        "1. 使い方 (このページ)",
        "2. 🟡 入力 ← ここだけ書き換える！黄色セル全部",
        "3. 🎯 結論 ← 結果を見る",
        "4. 📊 年次推移 ← 何歳でいくら持てるかグラフ",
        "5. 🔄 FIRE試算 ← 別シナリオの比較",
        "6. 🐪 らくだ先生について",
        "",
        "■ サイドFIREとは？",
        "・フルFIRE = 資産収入だけで生活費が全部賄える状態",
        "・サイドFIRE = 資産収入 + 緩めの労働(副業10-20万) で生活",
        "→ フルFIREより早く現実的に到達できる戦略",
        "",
        "■ 4%ルール (計算の根拠)",
        "・必要資産 = 年生活費 × 25",
        "  例: 年300万生活なら 300×25 = 7,500万円 必要 (フルFIRE)",
        "・サイドFIRE = (年生活費 - 副業年収) × 25",
        "  例: 副業年100万なら (300-100)×25 = 5,000万円 で到達",
        "",
        "■ 使い方（3ステップ）",
        "1. 「入力」シートで 🟡黄色セル を全部埋める (10〜12項目だけ)",
        "2. 「結論」シートを開く",
        "3. 🎯 サイドFIRE実現可否 / 不足額 / 達成年齢 を確認",
        "",
        "■ 値の変更でシミュレーション",
        "・月積立を5万増やすと何年早まる？ → 入力シートで月積立を増やして結論を見る",
        "・想定年利を 3% → 5% にすると？ → 同じく入力で変えて結論を見る",
        "",
        "■ 注意",
        "→ サンプル数字は伊神さんの実例ベース",
        "→ 想定年利は過去実績(オルカン+S&P500で 5-7%)。元本割れリスクあり",
        "→ 4%ルールは米国実績。日本では 3-3.5% 推奨説もあり",
    ]
    for i, line in enumerate(lines, start=3):
        c = ws.cell(row=i, column=2, value=line)
        if line.startswith("■"):
            c.font = F(size=13, bold=True, color=BRAND_BROWN)
        elif line.startswith("→"):
            c.font = SMALL
        else:
            c.font = NORMAL
        c.alignment = LEFT

    # ============ 入力 (全集約) ============
    ws = wb.create_sheet("入力")
    ws.sheet_properties.tabColor = INPUT_YELLOW
    for letter, w in zip("ABCDE", [4, 28, 20, 50, 4]):
        ws.column_dimensions[letter].width = w
    title_row(ws, "入力｜🟡黄色セルだけ書き換える", "全12項目・5分で埋まる")

    row = 4
    section(ws, row, "👤 基本情報"); row += 1
    input_field(ws, row, "現在年齢", 27, "0\"歳\"", "(2人のうち年上の方の年齢)"); CURRENT_AGE = row; row += 1
    input_field(ws, row, "目標FIRE達成年齢", 45, "0\"歳\"", "(何歳でセミリタイアしたいか)"); TARGET_AGE = row; row += 2

    section(ws, row, "💰 あなた（メイン稼ぎ手）の収支"); row += 1
    input_field(ws, row, "月手取り給料", 250000, "#,##0\"円\""); A_INCOME = row; row += 1
    input_field(ws, row, "ボーナス年合計（手取り）", 1000000, "#,##0\"円\""); A_BONUS = row; row += 1
    input_field(ws, row, "月生活費（あなたの分）", 100000, "#,##0\"円\""); A_LIVING = row; row += 2

    section(ws, row, "💑 パートナーの収支"); row += 1
    input_field(ws, row, "月手取り給料", 200000, "#,##0\"円\""); B_INCOME = row; row += 1
    input_field(ws, row, "ボーナス年合計（手取り）", 600000, "#,##0\"円\""); B_BONUS = row; row += 1
    input_field(ws, row, "月生活費（パートナーの分）", 80000, "#,##0\"円\""); B_LIVING = row; row += 2

    section(ws, row, "💎 現在の資産と投資設定（夫婦合算）"); row += 1
    input_field(ws, row, "現在の総資産（貯金+運用）", 2000000, "#,##0\"円\"", "(2人合計)"); CUR_ASSETS = row; row += 1
    input_field(ws, row, "月積立額（NISA+iDeCo合計）", 150000, "#,##0\"円\"", "(2人合計)"); MONTHLY_INVEST = row; row += 1
    input_field(ws, row, "想定年利", 0.05, "0.0%", "(オルカン+SP500なら5-6%程度)"); RATE = row; row += 2

    section(ws, row, "🎯 サイドFIRE後の想定"); row += 1
    input_field(ws, row, "FIRE後の月生活費（夫婦）", 250000, "#,##0\"円\"", "(リタイア後の家計)"); FIRE_LIVING = row; row += 1
    input_field(ws, row, "FIRE後の副業月収（緩い労働）", 100000, "#,##0\"円\"", "(資産取崩しを抑える分)"); FIRE_SIDE = row

    # 名前定義 (式から参照しやすく)
    nm = {
        "current_age": f"入力!$C${CURRENT_AGE}",
        "target_age": f"入力!$C${TARGET_AGE}",
        "a_income": f"入力!$C${A_INCOME}",
        "a_bonus": f"入力!$C${A_BONUS}",
        "a_living": f"入力!$C${A_LIVING}",
        "b_income": f"入力!$C${B_INCOME}",
        "b_bonus": f"入力!$C${B_BONUS}",
        "b_living": f"入力!$C${B_LIVING}",
        "cur_assets": f"入力!$C${CUR_ASSETS}",
        "monthly_invest": f"入力!$C${MONTHLY_INVEST}",
        "rate": f"入力!$C${RATE}",
        "fire_living": f"入力!$C${FIRE_LIVING}",
        "fire_side": f"入力!$C${FIRE_SIDE}",
    }

    # ============ 結論 (最重要シート) ============
    ws = wb.create_sheet("結論")
    ws.sheet_properties.tabColor = OUTPUT_GREEN
    for letter, w in zip("ABCDE", [4, 32, 22, 50, 4]):
        ws.column_dimensions[letter].width = w
    title_row(ws, "🎯 結論｜あなたのサイドFIRE達成見込み",
              "下の緑セルが答え（入力を変えるとリアルタイムで更新）")

    row = 4
    section(ws, row, "Step 1: 必要な総資産はいくら？（4%ルール）"); row += 1
    calc_field(ws, row, "FIRE後の年間生活費（夫婦）", f"={nm['fire_living']}*12", "#,##0\"円\""); FIRE_YEAR_LIVING = row; row += 1
    calc_field(ws, row, "FIRE後の年間副業収入", f"={nm['fire_side']}*12", "#,##0\"円\""); FIRE_YEAR_SIDE = row; row += 1
    calc_field(ws, row, "資産取崩しが必要な年額", f"=C{FIRE_YEAR_LIVING}-C{FIRE_YEAR_SIDE}", "#,##0\"円\"", "= 生活費 - 副業"); NEED_ANNUAL = row; row += 1
    result_field(ws, row, "💡 サイドFIREに必要な総資産", f"=C{NEED_ANNUAL}*25", "#,##0\"円\"", "(4%ルールで25倍)"); NEED_TOTAL = row; row += 2

    section(ws, row, "Step 2: 目標年齢に到達できる資産は？"); row += 1
    calc_field(ws, row, "目標FIREまでの年数", f"={nm['target_age']}-{nm['current_age']}", "0\"年\""); YEARS_LEFT = row; row += 1
    calc_field(ws, row, "夫婦の月積立額", f"={nm['monthly_invest']}", "#,##0\"円\""); row += 1
    calc_field(ws, row, "想定年利", f"={nm['rate']}", "0.0%"); row += 1
    result_field(ws, row, "💰 目標年齢時の予想資産", f"=FV({nm['rate']}/12,C{YEARS_LEFT}*12,-{nm['monthly_invest']},-{nm['cur_assets']})", "#,##0\"円\"", "(複利計算)"); ACTUAL_AT_TARGET = row; row += 2

    section(ws, row, "Step 3: サイドFIRE達成見込みは？"); row += 1
    result_field(ws, row, "✅ 達成判定", f'=IF(C{ACTUAL_AT_TARGET}>=C{NEED_TOTAL},"✓ 達成見込み！","✗ あと少し")', "@"); JUDGE = row; row += 1
    result_field(ws, row, "📊 達成率", f"=C{ACTUAL_AT_TARGET}/C{NEED_TOTAL}", "0.0%"); RATIO = row; row += 1
    result_field(ws, row, "📉 不足額（達成のためにあといくら必要）", f'=MAX(0,C{NEED_TOTAL}-C{ACTUAL_AT_TARGET})', "#,##0\"円\""); SHORTAGE = row; row += 1
    result_field(ws, row, "📈 月積立をあといくら増やせば達成？", f'=IFERROR(MAX(0,(PMT({nm["rate"]}/12,C{YEARS_LEFT}*12,-{nm["cur_assets"]},-C{NEED_TOTAL})-(-{nm["monthly_invest"]}))*-1),0)', "#,##0\"円\""); MORE_NEEDED = row; row += 2

    section(ws, row, "💡 別シナリオ（参考）"); row += 1
    calc_field(ws, row, "現状ペース継続 → 必要額到達は何歳？",
               f'=IFERROR({nm["current_age"]}+ROUND(NPER({nm["rate"]}/12,-{nm["monthly_invest"]},-{nm["cur_assets"]},C{NEED_TOTAL})/12,0),"-")',
               "0\"歳\"", "(月積立を続けて何歳で必要額に届くか)"); row += 1
    calc_field(ws, row, "今すぐ取崩したら何年もつか",
               f'=IFERROR(ROUND(NPER({nm["rate"]}/12,-C{NEED_ANNUAL}/12,-{nm["cur_assets"]})/12,1),"資産不足")',
               "0.0\"年\"", "(資産取崩し開始時の寿命)"); row += 2

    section(ws, row, "👀 入力シートで試すこと"); row += 1
    for text in [
        "・月積立を 15万 → 20万 にすると達成年齢どう変わる？",
        "・想定年利を 5% → 4% に下げると不足額は？",
        "・FIRE後の副業を 10万 → 15万 にすると必要資産がどれだけ減る？",
        "・目標年齢を 45 → 50歳 にすると余裕は出る？",
    ]:
        c = ws.cell(row=row, column=2, value=text)
        c.font = NORMAL
        c.alignment = LEFT
        row += 1

    # ============ 年次推移 ============
    ws = wb.create_sheet("年次推移")
    ws.sheet_properties.tabColor = "BBDEFB"
    for letter, w in zip("ABCDE", [4, 10, 22, 22, 28]):
        ws.column_dimensions[letter].width = w
    title_row(ws, "📊 年次推移｜年齢別の予想資産")
    row = 4
    section(ws, row, "現在の入力ペースで何歳でいくら？"); row += 1
    headers = ["年齢", "経過年数", "予想資産", "必要資産との差"]
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=2 + i, value=h)
        c.font = H1
        c.fill = FILL_HEADER
        c.alignment = CENTER
        c.border = BORDER_THIN
    row += 1
    start_row = row
    for n in range(0, 41):
        # 年齢
        ws.cell(row=row, column=2, value=f"={nm['current_age']}+{n}").number_format = "0\"歳\""
        ws.cell(row=row, column=2).alignment = CENTER
        ws.cell(row=row, column=2).border = BORDER_THIN
        # 経過年数
        ws.cell(row=row, column=3, value=n).number_format = "0\"年後\""
        ws.cell(row=row, column=3).alignment = CENTER
        ws.cell(row=row, column=3).border = BORDER_THIN
        # 予想資産
        if n == 0:
            ws.cell(row=row, column=4, value=f"={nm['cur_assets']}")
        else:
            ws.cell(row=row, column=4, value=f"=FV({nm['rate']}/12,{n}*12,-{nm['monthly_invest']},-{nm['cur_assets']})")
        ws.cell(row=row, column=4).number_format = "#,##0\"円\""
        ws.cell(row=row, column=4).alignment = RIGHT
        ws.cell(row=row, column=4).border = BORDER_THIN
        # 必要資産との差
        ws.cell(row=row, column=5, value=f"=D{row}-結論!C{NEED_TOTAL}")
        ws.cell(row=row, column=5).number_format = "#,##0\"円\";[Red]-#,##0\"円\""
        ws.cell(row=row, column=5).alignment = RIGHT
        ws.cell(row=row, column=5).border = BORDER_THIN
        row += 1

    # ============ FIRE試算 (シナリオ比較) ============
    ws = wb.create_sheet("FIRE試算")
    ws.sheet_properties.tabColor = "D1C4E9"
    for letter, w in zip("ABCDEF", [4, 22, 16, 16, 16, 30]):
        ws.column_dimensions[letter].width = w
    title_row(ws, "🔄 FIRE試算｜シナリオ比較", "月積立/年利を変えて達成資産がどう変わるか")
    row = 4
    section(ws, row, "5パターン比較（10年後の資産）", span=6); row += 1
    headers = ["シナリオ", "月積立", "年利", "期間(年)", "予想資産", "備考"]
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=1 + i, value=h)
        c.font = H1
        c.fill = FILL_HEADER
        c.alignment = CENTER
        c.border = BORDER_THIN
    row += 1
    scenarios = [
        ("①現状ペース", "B", 10, "C", "あなたの入力値そのまま"),
        ("②積立 +5万", 50000, 10, "C", "月積立を5万増やす"),
        ("③年利 +1%", "B", 10, 0.01, "投資配分をやや攻めに"),
        ("④目標年齢 +5年", "B", 15, "C", "5年長く働く"),
        ("⑤積立0で運用のみ", 0, 10, "C", "今ある資産だけ運用"),
    ]
    for s_name, invest_mod, years, rate_mod, memo in scenarios:
        ws.cell(row=row, column=1, value=s_name).font = LABEL
        ws.cell(row=row, column=1).border = BORDER_THIN
        # 月積立
        if invest_mod == "B":
            invest_formula = f"={nm['monthly_invest']}"
        elif isinstance(invest_mod, int):
            if invest_mod == 0:
                invest_formula = 0
            else:
                invest_formula = f"={nm['monthly_invest']}+{invest_mod}"
        ws.cell(row=row, column=2, value=invest_formula).number_format = "#,##0\"円\""
        ws.cell(row=row, column=2).alignment = RIGHT
        ws.cell(row=row, column=2).border = BORDER_THIN
        # 年利
        if rate_mod == "C":
            rate_formula = f"={nm['rate']}"
        else:
            rate_formula = f"={nm['rate']}+{rate_mod}"
        ws.cell(row=row, column=3, value=rate_formula).number_format = "0.0%"
        ws.cell(row=row, column=3).alignment = CENTER
        ws.cell(row=row, column=3).border = BORDER_THIN
        # 期間
        ws.cell(row=row, column=4, value=years).number_format = "0\"年\""
        ws.cell(row=row, column=4).alignment = CENTER
        ws.cell(row=row, column=4).border = BORDER_THIN
        # 予想資産 (FV)
        ws.cell(row=row, column=5, value=f"=FV(C{row}/12,D{row}*12,-B{row},-{nm['cur_assets']})")
        ws.cell(row=row, column=5).number_format = "#,##0\"円\""
        ws.cell(row=row, column=5).alignment = RIGHT
        ws.cell(row=row, column=5).fill = FILL_OUTPUT
        ws.cell(row=row, column=5).font = LABEL
        ws.cell(row=row, column=5).border = BORDER_OUTPUT
        # 備考
        ws.cell(row=row, column=6, value=memo).font = SMALL
        ws.cell(row=row, column=6).alignment = LEFT
        ws.cell(row=row, column=6).border = BORDER_THIN
        row += 1

    # ============ らくだ先生について ============
    ws = wb.create_sheet("らくだ先生について")
    ws.sheet_properties.tabColor = BRAND_BEIGE
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 95
    info = [
        ("🐪", TITLE, "残業嫌いのらくだ先生"),
        ("", NORMAL, ""),
        ("■", H2, "プロフィール"),
        ("", NORMAL, "公立中学校教員（国語・20代）"),
        ("", NORMAL, "ICT・AIで残業ゼロ ｜ 副業10万 ｜ 5年で2000万円貯金"),
        ("", NORMAL, "日本株デイトレで-50万やった失敗から、低リスクインデックス投資に振り切り。"),
        ("", NORMAL, ""),
        ("■", H2, "発信中"),
        ("", NORMAL, "X (Twitter): @rakuda_sensei"),
        ("", NORMAL, "note: https://note.com/rakuda_sensei"),
        ("", NORMAL, "BOOTH: https://rakuda-sensei.booth.pm"),
        ("", NORMAL, ""),
        ("■", H2, "ご利用について"),
        ("", NORMAL, "・本商品の個人利用は自由"),
        ("", NORMAL, "・二次配布・転売は禁止"),
        ("", NORMAL, "・本商品の数式・テンプレで生じた結果について作者は責任を負いません"),
        ("", NORMAL, "・改善要望・不具合報告は X DM までお気軽に"),
    ]
    for i, (icon, font, text) in enumerate(info, start=1):
        ws.cell(row=i, column=1, value=icon)
        c = ws.cell(row=i, column=2, value=text)
        c.font = font
        c.alignment = LEFT

    out = OUT_DIR / "夫婦で実践｜サイドFIRE計画シート.xlsx"
    wb.save(out)
    return out


if __name__ == "__main__":
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out = build()
    print(f"✅ 完成: {out.name} ({out.stat().st_size // 1024}KB)")
