#!/usr/bin/env python3
"""
サイドFIRE計画シート v4.1
- 生活費内訳化 + 貯金/運用分離 + 結論ビジュアル化
- 値の右に「〇万円/〇億円」読みやすい表示
- Step 5「ここを動かす」を自動計算化
- 何年もつ計算のサイン修正（永久にも対応）
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[3]
OUT_DIR = ROOT / "projects" / "rakuda-sensei" / "products" / "digital"

# ====== カラー ======
BRAND_BROWN = "8B6F47"
BRAND_BEIGE = "E8D5A7"
BRAND_DARK = "3D2F1F"
INPUT_YELLOW = "FFF59D"
OUTPUT_GREEN = "A5D6A7"
HERO_GREEN = "66BB6A"
WARNING_RED = "EF5350"
SAVINGS_BLUE = "B3E5FC"
INVEST_PURPLE = "D1C4E9"
ACCENT_ORANGE = "FFB74D"
MAN_GRAY = "FAFAFA"


def F(size=11, bold=False, color=BRAND_DARK, italic=False):
    return Font(name="メイリオ", size=size, bold=bold, color=color, italic=italic)

TITLE = F(size=22, bold=True)
HERO = F(size=28, bold=True, color="FFFFFF")
H1 = F(size=13, bold=True, color="FFFFFF")
H2 = F(size=13, bold=True, color=BRAND_BROWN)
H3 = F(size=11, bold=True, color=BRAND_BROWN)
LABEL = F(size=11, bold=True)
NORMAL = F(size=11)
SMALL = F(size=10, color="666666")
BIG = F(size=18, bold=True, color=BRAND_DARK)
HUGE = F(size=36, bold=True, color="FFFFFF")
NOTE = F(size=9, italic=True, color="999999")
MARKER_RED = F(size=12, bold=True, color="D32F2F")
MAN = F(size=11, bold=True, color="555555")
MAN_BIG = F(size=14, bold=True, color="2E7D32")

FILL_HEADER = PatternFill(start_color=BRAND_BROWN, end_color=BRAND_BROWN, fill_type="solid")
FILL_SECTION = PatternFill(start_color=BRAND_BEIGE, end_color=BRAND_BEIGE, fill_type="solid")
FILL_INPUT = PatternFill(start_color=INPUT_YELLOW, end_color=INPUT_YELLOW, fill_type="solid")
FILL_OUTPUT = PatternFill(start_color=OUTPUT_GREEN, end_color=OUTPUT_GREEN, fill_type="solid")
FILL_HERO = PatternFill(start_color=HERO_GREEN, end_color=HERO_GREEN, fill_type="solid")
FILL_WARN = PatternFill(start_color=WARNING_RED, end_color=WARNING_RED, fill_type="solid")
FILL_SAVINGS = PatternFill(start_color=SAVINGS_BLUE, end_color=SAVINGS_BLUE, fill_type="solid")
FILL_INVEST = PatternFill(start_color=INVEST_PURPLE, end_color=INVEST_PURPLE, fill_type="solid")
FILL_ACCENT = PatternFill(start_color=ACCENT_ORANGE, end_color=ACCENT_ORANGE, fill_type="solid")
FILL_MAN = PatternFill(start_color=MAN_GRAY, end_color=MAN_GRAY, fill_type="solid")

T = Side(border_style="thin", color="CCCCCC")
M = Side(border_style="medium", color=BRAND_BROWN)
M_RED = Side(border_style="medium", color="D32F2F")
B_THIN = Border(left=T, right=T, top=T, bottom=T)
B_BOLD = Border(left=M, right=M, top=M, bottom=M)
B_HERO = Border(left=M_RED, right=M_RED, top=M_RED, bottom=M_RED)

CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
RGT = Alignment(horizontal="right", vertical="center")
LFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def man_formula(value_ref: str) -> str:
    """円の値を『X億X,XXX万円』『X,XXX万円』形式に変換するExcel式。"""
    return (
        f'=IF(NOT(ISNUMBER({value_ref})),"",'
        f'IF(ABS({value_ref})>=100000000,'
        f'TEXT(INT({value_ref}/100000000),"0")&"億"&'
        f'IF(MOD({value_ref},100000000)>=10000,TEXT(ROUND(MOD({value_ref},100000000)/10000,0),"#,##0")&"万円","円"),'
        f'IF(ABS({value_ref})>=10000,TEXT(ROUND({value_ref}/10000,0),"#,##0")&"万円",'
        f'TEXT({value_ref},"#,##0")&"円")))'
    )


def title_row(ws, title: str, sub: str = ""):
    ws.cell(row=1, column=1, value="🐪").font = F(size=24)
    c = ws.cell(row=1, column=2, value=title); c.font = TITLE
    ws.row_dimensions[1].height = 36
    if sub:
        ws.cell(row=2, column=2, value=sub).font = SMALL


def section_bar(ws, row: int, text: str, span: int = 6):
    for c in range(1, span + 1):
        ws.cell(row=row, column=c).fill = FILL_SECTION
        ws.cell(row=row, column=c).border = B_THIN
    cell = ws.cell(row=row, column=1, value=text); cell.font = H2; cell.alignment = LFT
    ws.row_dimensions[row].height = 24


def field_input(ws, row: int, label: str, value, fmt: str = "#,##0", note: str = "", show_man: bool = True):
    ws.cell(row=row, column=2, value=label).font = LABEL
    ws.cell(row=row, column=2).alignment = LFT
    ws.cell(row=row, column=2).border = B_THIN
    ic = ws.cell(row=row, column=3, value=value)
    ic.fill = FILL_INPUT; ic.border = B_BOLD; ic.font = F(size=12, bold=True)
    ic.alignment = RGT; ic.number_format = fmt
    if show_man:
        mc = ws.cell(row=row, column=4, value=man_formula(f"C{row}"))
        mc.font = MAN; mc.alignment = RGT; mc.fill = FILL_MAN; mc.border = B_THIN
    if note:
        ws.cell(row=row, column=5, value=note).font = NOTE; ws.cell(row=row, column=5).alignment = LFT


def field_calc(ws, row: int, label: str, formula, fmt: str = "#,##0", note: str = "", show_man: bool = True):
    ws.cell(row=row, column=2, value=label).font = NORMAL
    ws.cell(row=row, column=2).alignment = LFT
    ws.cell(row=row, column=2).border = B_THIN
    cc = ws.cell(row=row, column=3, value=formula)
    cc.border = B_THIN; cc.font = NORMAL; cc.alignment = RGT; cc.number_format = fmt
    if show_man:
        mc = ws.cell(row=row, column=4, value=man_formula(f"C{row}"))
        mc.font = MAN; mc.alignment = RGT; mc.fill = FILL_MAN; mc.border = B_THIN
    if note:
        ws.cell(row=row, column=5, value=note).font = SMALL; ws.cell(row=row, column=5).alignment = LFT


def field_result(ws, row: int, label: str, formula, fmt: str = "#,##0", note: str = "", show_man: bool = True):
    ws.cell(row=row, column=2, value=label).font = LABEL
    ws.cell(row=row, column=2).alignment = LFT
    ws.cell(row=row, column=2).border = B_BOLD
    rc = ws.cell(row=row, column=3, value=formula)
    rc.fill = FILL_OUTPUT; rc.border = B_BOLD; rc.font = BIG
    rc.alignment = RGT; rc.number_format = fmt
    if show_man:
        mc = ws.cell(row=row, column=4, value=man_formula(f"C{row}"))
        mc.font = MAN_BIG; mc.alignment = RGT; mc.fill = FILL_MAN; mc.border = B_BOLD
    ws.row_dimensions[row].height = 28
    if note:
        ws.cell(row=row, column=5, value=note).font = SMALL; ws.cell(row=row, column=5).alignment = LFT


# =========================================================================
def build():
    wb = Workbook()
    wb.remove(wb.active)

    # =================== 使い方 ===================
    ws = wb.create_sheet("使い方", 0)
    ws.sheet_properties.tabColor = BRAND_BROWN
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 95
    title_row(ws, "夫婦で実践｜サイドFIRE計画シート v4.1",
              "貯金と運用を分離・生活費は内訳から自動・読みやすい〇万円表示付き")
    lines = [
        "",
        "■ 読む順番",
        "①使い方 (このページ)",
        "②生活費 ← 月の支出を項目別に入力",
        "③入力 ← 黄色セル全部書き換える（年齢・収入・資産・投資設定）",
        "④🎯 結論 ← ここを開けば全部わかる",
        "⑤年次推移 ← 何歳でいくら持てるか",
        "⑥FIRE試算 ← 別シナリオ比較",
        "",
        "■ 「〇万円」表示について",
        "・大きい桁の数字（億・千万）は右隣のセルに『〇億〇万円』形式で表示されます",
        "・例: 30,000,000円 → 「3,000万円」",
        "・例: 150,000,000円 → 「1億5,000万円」",
        "",
        "■ サイドFIREの計算根拠",
        "・必要資産 = 年取崩額 × 25倍（4%ルール）",
        "・年取崩額 = FIRE後の年生活費 − 副業年収",
        "・例: 年300万生活で年100万副業なら (300-100)×25 = 5,000万必要",
        "",
        "■ 注意",
        "→ 数字はサンプルです。あなたの状況に書き換えてください",
        "→ 想定年利は過去実績(オルカン+SP500で5-6%)、元本割れリスクあり",
        "→ 4%ルールは米国実績、日本では3-3.5%推奨説あり",
    ]
    for i, line in enumerate(lines, start=3):
        c = ws.cell(row=i, column=2, value=line)
        if line.startswith("■"):
            c.font = H2
        elif line.startswith("→"):
            c.font = SMALL
        else:
            c.font = NORMAL
        c.alignment = LFT

    # =================== 生活費 ===================
    ws = wb.create_sheet("生活費")
    ws.sheet_properties.tabColor = ACCENT_ORANGE
    for letter, w in zip("ABCDEF", [4, 24, 16, 14, 4, 36]):
        ws.column_dimensions[letter].width = w
    title_row(ws, "生活費｜月の家計内訳", "🟡黄色セルに金額入力 → 合計が『入力』シートに自動反映 / 右に万円表示")

    row = 4
    section_bar(ws, row, "📌 今の月の生活費（現在）", span=6); row += 1
    # ヘッダ
    for col_idx, (col, text) in enumerate([("B", "項目"), ("C", "月額"), ("D", "万円表示"), ("F", "備考")]):
        c = ws.cell(row=row, column=ord(col) - 64, value=text)
        c.font = H1; c.fill = FILL_HEADER; c.alignment = CTR; c.border = B_THIN
    row += 1
    current_items = [
        ("住居費", 90000, "家賃・住宅ローン・管理費"),
        ("食費", 80000, "自炊中心＋外食少し"),
        ("水道光熱費", 22000, "電気・ガス・水道"),
        ("通信費", 18000, "スマホ＋ネット"),
        ("日用品", 12000, "消耗品・雑貨"),
        ("交通費", 20000, "車・公共交通"),
        ("医療・保険", 15000, "医療費＋保険料"),
        ("娯楽・交際費", 25000, "レジャー・外食"),
        ("社会保険料", 40000, "国保・年金等"),
        ("教育費", 0, "子どもがいる場合のみ"),
        ("その他", 0, "雑費"),
    ]
    current_start = row
    for label, amount, memo in current_items:
        ws.cell(row=row, column=2, value=label).font = NORMAL
        ws.cell(row=row, column=2).alignment = LFT; ws.cell(row=row, column=2).border = B_THIN
        ic = ws.cell(row=row, column=3, value=amount)
        ic.fill = FILL_INPUT; ic.border = B_BOLD; ic.font = LABEL
        ic.alignment = RGT; ic.number_format = "#,##0\"円\""
        mc = ws.cell(row=row, column=4, value=man_formula(f"C{row}"))
        mc.font = MAN; mc.alignment = RGT; mc.fill = FILL_MAN; mc.border = B_THIN
        ws.cell(row=row, column=6, value=memo).font = SMALL; ws.cell(row=row, column=6).alignment = LFT; ws.cell(row=row, column=6).border = B_THIN
        row += 1
    current_end = row - 1
    # 合計
    ws.cell(row=row, column=2, value="🔢 合計（月）").font = LABEL
    ws.cell(row=row, column=2).fill = FILL_OUTPUT; ws.cell(row=row, column=2).border = B_BOLD; ws.cell(row=row, column=2).alignment = LFT
    rc = ws.cell(row=row, column=3, value=f"=SUM(C{current_start}:C{current_end})")
    rc.fill = FILL_OUTPUT; rc.border = B_BOLD; rc.font = BIG
    rc.alignment = RGT; rc.number_format = "#,##0\"円\""
    mc = ws.cell(row=row, column=4, value=man_formula(f"C{row}"))
    mc.font = MAN_BIG; mc.alignment = RGT; mc.fill = FILL_OUTPUT; mc.border = B_BOLD
    ws.row_dimensions[row].height = 28
    CURRENT_TOTAL = row
    ws.cell(row=row, column=6, value="← 『入力』シートに自動反映").font = MARKER_RED
    row += 2

    section_bar(ws, row, "📌 FIRE後の想定生活費（仕事減らした後）", span=6); row += 1
    for col_idx, (col, text) in enumerate([("B", "項目"), ("C", "月額"), ("D", "万円表示"), ("F", "変化")]):
        c = ws.cell(row=row, column=ord(col) - 64, value=text)
        c.font = H1; c.fill = FILL_HEADER; c.alignment = CTR; c.border = B_THIN
    row += 1
    future_items = [
        ("住居費", 90000, "ローン完済なら下がる"),
        ("食費", 70000, "在宅で少し下がる"),
        ("水道光熱費", 25000, "在宅で少し上がる"),
        ("通信費", 18000, "変わらず"),
        ("日用品", 12000, "変わらず"),
        ("交通費", 15000, "通勤費減"),
        ("医療・保険", 30000, "国保＋年齢上昇"),
        ("娯楽・交際費", 35000, "時間できて増える"),
        ("社会保険料", 0, "副業のみなら大幅減"),
        ("教育費", 30000, "子供想定（任意）"),
        ("その他", 5000, "趣味等"),
    ]
    future_start = row
    for label, amount, memo in future_items:
        ws.cell(row=row, column=2, value=label).font = NORMAL
        ws.cell(row=row, column=2).alignment = LFT; ws.cell(row=row, column=2).border = B_THIN
        ic = ws.cell(row=row, column=3, value=amount)
        ic.fill = FILL_INPUT; ic.border = B_BOLD; ic.font = LABEL
        ic.alignment = RGT; ic.number_format = "#,##0\"円\""
        mc = ws.cell(row=row, column=4, value=man_formula(f"C{row}"))
        mc.font = MAN; mc.alignment = RGT; mc.fill = FILL_MAN; mc.border = B_THIN
        ws.cell(row=row, column=6, value=memo).font = SMALL; ws.cell(row=row, column=6).alignment = LFT; ws.cell(row=row, column=6).border = B_THIN
        row += 1
    future_end = row - 1
    ws.cell(row=row, column=2, value="🔢 FIRE後合計（月）").font = LABEL
    ws.cell(row=row, column=2).fill = FILL_OUTPUT; ws.cell(row=row, column=2).border = B_BOLD; ws.cell(row=row, column=2).alignment = LFT
    rc = ws.cell(row=row, column=3, value=f"=SUM(C{future_start}:C{future_end})")
    rc.fill = FILL_OUTPUT; rc.border = B_BOLD; rc.font = BIG
    rc.alignment = RGT; rc.number_format = "#,##0\"円\""
    mc = ws.cell(row=row, column=4, value=man_formula(f"C{row}"))
    mc.font = MAN_BIG; mc.alignment = RGT; mc.fill = FILL_OUTPUT; mc.border = B_BOLD
    ws.row_dimensions[row].height = 28
    FUTURE_TOTAL = row
    ws.cell(row=row, column=6, value="← 『入力』シートに自動反映").font = MARKER_RED

    # =================== 入力 ===================
    ws = wb.create_sheet("入力")
    ws.sheet_properties.tabColor = INPUT_YELLOW
    for letter, w in zip("ABCDEF", [4, 30, 18, 18, 40, 4]):
        ws.column_dimensions[letter].width = w
    title_row(ws, "入力｜🟡黄色セルだけ書き換える", "全11項目・5分で埋まる / 値の右に万円表示")

    row = 4
    section_bar(ws, row, "👤 基本情報"); row += 1
    field_input(ws, row, "現在年齢", 27, "0\"歳\"", "(夫婦のうち年上の方)", show_man=False); CURRENT_AGE = row; row += 1
    field_input(ws, row, "目標 FIRE 年齢", 45, "0\"歳\"", "(何歳でセミリタイア？)", show_man=False); TARGET_AGE = row; row += 2

    section_bar(ws, row, "💰 夫婦の月収入"); row += 1
    field_input(ws, row, "夫婦の月手取り合計", 450000, "#,##0\"円\"", "(2人の手取り合計)"); MONTH_INCOME = row; row += 1
    field_input(ws, row, "夫婦のボーナス年合計", 1600000, "#,##0\"円\"", "(2人の年ボーナス手取り合計)"); BONUS = row; row += 1
    field_calc(ws, row, "🔢 年手取り合計", f"=C{MONTH_INCOME}*12+C{BONUS}", "#,##0\"円\"", "(自動)"); YEAR_INCOME = row; row += 2

    section_bar(ws, row, "💸 月の生活費（『生活費』シートから自動反映）"); row += 1
    field_calc(ws, row, "月生活費（現在）", f"=生活費!C{CURRENT_TOTAL}", "#,##0\"円\"", "← 編集は『生活費』シートで"); CUR_LIVING = row; row += 1
    field_calc(ws, row, "🔢 月貯蓄可能額", f"=C{MONTH_INCOME}-C{CUR_LIVING}", "#,##0\"円\"", "(月収 − 月生活費)"); MONTH_SAVABLE = row; row += 1
    field_calc(ws, row, "🔢 年貯蓄可能額（ボーナス込）", f"=C{MONTH_SAVABLE}*12+C{BONUS}", "#,##0\"円\""); YEAR_SAVABLE = row; row += 2

    section_bar(ws, row, "🏦 今ある資産（貯金と運用を分けて入力）"); row += 1
    field_input(ws, row, "現在の貯金（現金・利率0%）", 500000, "#,##0\"円\"", "(銀行預金など)"); CUR_CASH = row
    ws.cell(row=row, column=3).fill = FILL_SAVINGS
    ws.cell(row=row, column=2).font = F(size=11, bold=True, color="0277BD")
    row += 1
    field_input(ws, row, "現在の運用資産（投資中・年利で成長）", 1500000, "#,##0\"円\"", "(NISA・iDeCo等の投信)"); CUR_INVEST = row
    ws.cell(row=row, column=3).fill = FILL_INVEST
    ws.cell(row=row, column=2).font = F(size=11, bold=True, color="6A1B9A")
    row += 1
    field_calc(ws, row, "🔢 総資産（現在）", f"=C{CUR_CASH}+C{CUR_INVEST}", "#,##0\"円\""); CUR_TOTAL = row; row += 2

    section_bar(ws, row, "📈 投資設定"); row += 1
    field_input(ws, row, "月積立額（投資への積立）", 150000, "#,##0\"円\"", "(NISA+iDeCo合計)"); MONTHLY_INVEST = row; row += 1
    field_input(ws, row, "想定年利", 0.05, "0.0%", "(オルカン+SP500で5-6%)", show_man=False); RATE = row; row += 2

    section_bar(ws, row, "🎯 サイドFIRE後の想定"); row += 1
    field_calc(ws, row, "FIRE後の月生活費（『生活費』シート参照）", f"=生活費!C{FUTURE_TOTAL}", "#,##0\"円\"", "← 編集は『生活費』シートで"); FIRE_LIVING = row; row += 1
    field_input(ws, row, "FIRE後の月副業収入", 100000, "#,##0\"円\"", "(緩い労働で稼ぐ分)"); FIRE_SIDE = row

    # 名前定義
    INPUT = "入力!"
    nm = {
        "current_age": f"{INPUT}$C${CURRENT_AGE}",
        "target_age": f"{INPUT}$C${TARGET_AGE}",
        "month_income": f"{INPUT}$C${MONTH_INCOME}",
        "bonus": f"{INPUT}$C${BONUS}",
        "year_income": f"{INPUT}$C${YEAR_INCOME}",
        "cur_living": f"{INPUT}$C${CUR_LIVING}",
        "month_savable": f"{INPUT}$C${MONTH_SAVABLE}",
        "year_savable": f"{INPUT}$C${YEAR_SAVABLE}",
        "cur_cash": f"{INPUT}$C${CUR_CASH}",
        "cur_invest": f"{INPUT}$C${CUR_INVEST}",
        "cur_total": f"{INPUT}$C${CUR_TOTAL}",
        "monthly_invest": f"{INPUT}$C${MONTHLY_INVEST}",
        "rate": f"{INPUT}$C${RATE}",
        "fire_living": f"{INPUT}$C${FIRE_LIVING}",
        "fire_side": f"{INPUT}$C${FIRE_SIDE}",
    }

    # =================== 結論 ===================
    ws = wb.create_sheet("結論")
    ws.sheet_properties.tabColor = OUTPUT_GREEN
    for letter, w in zip("ABCDEF", [4, 30, 20, 18, 26, 4]):
        ws.column_dimensions[letter].width = w
    title_row(ws, "🎯 結論｜あなたのサイドFIRE達成診断",
              "上から順に読むだけで全部わかる / 値の右に万円表示")

    row = 4
    # ▼ヒーロー結果バナー▼
    ws.cell(row=row, column=2, value="🎉 サイドFIRE 達成可否").font = F(size=14, bold=True, color="FFFFFF")
    for c in range(2, 6):
        ws.cell(row=row, column=c).fill = FILL_HERO
        ws.cell(row=row, column=c).border = B_HERO
    ws.row_dimensions[row].height = 32
    HERO_LABEL_ROW = row; row += 1
    ws.merge_cells(start_row=row, end_row=row, start_column=2, end_column=5)
    judge_cell = ws.cell(row=row, column=2)
    judge_cell.fill = FILL_HERO; judge_cell.border = B_HERO
    judge_cell.font = HUGE; judge_cell.alignment = CTR
    judge_cell.value = None
    ws.row_dimensions[row].height = 60
    HERO_JUDGE_ROW = row; row += 2

    # ▼Step 1▼
    section_bar(ws, row, "Step 1: 必要な総資産を計算（4%ルール）"); row += 1
    field_calc(ws, row, "FIRE後の年生活費", f"={nm['fire_living']}*12", "#,##0\"円\""); FIRE_YEAR_LIVING = row; row += 1
    field_calc(ws, row, "FIRE後の年副業収入", f"={nm['fire_side']}*12", "#,##0\"円\""); FIRE_YEAR_SIDE = row; row += 1
    field_calc(ws, row, "資産取崩しが必要な年額", f"=C{FIRE_YEAR_LIVING}-C{FIRE_YEAR_SIDE}", "#,##0\"円\"", "= 年生活費 − 年副業"); NEED_ANNUAL = row; row += 1
    field_result(ws, row, "💡 サイドFIRE 必要総資産", f"=C{NEED_ANNUAL}*25", "#,##0\"円\"", "(年取崩額 × 25倍)"); NEED_TOTAL = row; row += 2

    # ▼Step 2▼
    section_bar(ws, row, "Step 2: 目標年齢時に達成できる資産"); row += 1
    field_calc(ws, row, "目標FIREまでの年数", f"={nm['target_age']}-{nm['current_age']}", "0\"年\"", show_man=False); YEARS_LEFT = row; row += 1
    field_calc(ws, row, "💧 目標年齢時の貯金（現金・成長なし）", f"={nm['cur_cash']}", "#,##0\"円\"", "(現金は複利成長しない前提)"); FUTURE_CASH = row
    ws.cell(row=row, column=3).fill = FILL_SAVINGS
    row += 1
    field_calc(ws, row, "📈 目標年齢時の運用資産（複利成長）",
               f"=FV({nm['rate']}/12,C{YEARS_LEFT}*12,-{nm['monthly_invest']},-{nm['cur_invest']})",
               "#,##0\"円\"", "(月積立を続けた前提)"); FUTURE_INVEST = row
    ws.cell(row=row, column=3).fill = FILL_INVEST
    row += 1
    field_result(ws, row, "💰 目標年齢時の予想総資産", f"=C{FUTURE_CASH}+C{FUTURE_INVEST}", "#,##0\"円\""); ACTUAL_AT_TARGET = row; row += 2

    # ▼Step 3▼
    section_bar(ws, row, "Step 3: 達成診断（数字で見る）"); row += 1
    field_result(ws, row, "📊 達成率", f"=C{ACTUAL_AT_TARGET}/C{NEED_TOTAL}", "0.0%", show_man=False); RATIO = row; row += 1
    field_result(ws, row, "📉 不足額（達成にあといくら必要？）", f"=MAX(0,C{NEED_TOTAL}-C{ACTUAL_AT_TARGET})", "#,##0\"円\""); SHORTAGE = row; row += 1
    field_result(ws, row, "📈 月積立をあといくら増やせば達成？",
                 f"=IFERROR(MAX(0,(PMT({nm['rate']}/12,C{YEARS_LEFT}*12,-{nm['cur_invest']},-(C{NEED_TOTAL}-C{FUTURE_CASH}))-(-{nm['monthly_invest']}))*-1),0)",
                 "#,##0\"円\""); MORE_INVEST = row; row += 2

    judge_cell.value = f'=IF(C{ACTUAL_AT_TARGET}>=C{NEED_TOTAL},"✓ 達成見込み！おめでとう🐪","✗ 達成にあと一歩")'

    # ▼Step 4▼
    section_bar(ws, row, "Step 4: 別の見方で確認"); row += 1
    # 達成年齢
    field_calc(ws, row, "現状ペース継続 → 必要額到達は何歳？",
               f'=IFERROR({nm["current_age"]}+ROUND(NPER({nm["rate"]}/12,-{nm["monthly_invest"]},-{nm["cur_invest"]},C{NEED_TOTAL}-{nm["cur_cash"]})/12,0)&"歳","計算不能")',
               "@", "(月積立を続けて何歳で達成？)", show_man=False); row += 1
    # 何年もつ（修正版・サイン正しい・永久対応）
    field_calc(ws, row, "達成資産で何年もつ？（運用継続前提）",
               f'=IF(C{NEED_ANNUAL}<=0,"永久に",'
               f'IF(C{ACTUAL_AT_TARGET}*{nm["rate"]}>=C{NEED_ANNUAL},"永久に減らない（運用益で賄える）",'
               f'IFERROR(ROUND(NPER({nm["rate"]}/12,-C{NEED_ANNUAL}/12,C{ACTUAL_AT_TARGET})/12,1)&"年","計算不能")))',
               "@", "(取崩し中も年利で運用続けた場合)", show_man=False); row += 1
    field_calc(ws, row, "FIRE達成までの月貯蓄ペース", f"={nm['month_savable']}", "#,##0\"円\"", "(月収−月生活費)"); row += 2

    # ▼Step 5: 自動計算式の提案▼
    section_bar(ws, row, "Step 5: ここを動かせば達成できる（自動計算）"); row += 1
    # ヘッダ
    ws.cell(row=row, column=2, value="変える項目").font = H1
    ws.cell(row=row, column=2).fill = FILL_HEADER; ws.cell(row=row, column=2).alignment = CTR; ws.cell(row=row, column=2).border = B_THIN
    ws.cell(row=row, column=3, value="現状").font = H1
    ws.cell(row=row, column=3).fill = FILL_HEADER; ws.cell(row=row, column=3).alignment = CTR; ws.cell(row=row, column=3).border = B_THIN
    ws.cell(row=row, column=4, value="→ 達成ライン").font = H1
    ws.cell(row=row, column=4).fill = FILL_HEADER; ws.cell(row=row, column=4).alignment = CTR; ws.cell(row=row, column=4).border = B_THIN
    ws.cell(row=row, column=5, value="一言メモ").font = H1
    ws.cell(row=row, column=5).fill = FILL_HEADER; ws.cell(row=row, column=5).alignment = CTR; ws.cell(row=row, column=5).border = B_THIN
    row += 1

    def step5_row(ws, row, label, current_formula, target_formula, memo):
        ws.cell(row=row, column=2, value=label).font = LABEL
        ws.cell(row=row, column=2).alignment = LFT; ws.cell(row=row, column=2).border = B_THIN
        a = ws.cell(row=row, column=3, value=current_formula)
        a.font = NORMAL; a.alignment = CTR; a.border = B_THIN
        b = ws.cell(row=row, column=4, value=target_formula)
        b.font = LABEL; b.alignment = CTR; b.border = B_BOLD; b.fill = FILL_OUTPUT
        ws.cell(row=row, column=5, value=memo).font = SMALL
        ws.cell(row=row, column=5).alignment = LFT; ws.cell(row=row, column=5).border = B_THIN
        ws.row_dimensions[row].height = 24

    # A. 月積立を増やす
    step5_row(ws, row,
        "A. 月積立を増やす",
        f'=TEXT(ROUND({nm["monthly_invest"]}/10000,1),"0.0")&"万円/月"',
        f'=IF(C{ACTUAL_AT_TARGET}>=C{NEED_TOTAL},"現状で達成済",TEXT(ROUND(({nm["monthly_invest"]}+C{MORE_INVEST})/10000,1),"0.0")&"万円/月")',
        "投資への月積立額"); row += 1

    # B. 目標年齢を遅らせる
    step5_row(ws, row,
        "B. FIRE開始年齢を遅らせる",
        f'={nm["target_age"]}&"歳"',
        f'=IF(C{ACTUAL_AT_TARGET}>=C{NEED_TOTAL},"現状で達成済",IFERROR(ROUND({nm["current_age"]}+NPER({nm["rate"]}/12,-{nm["monthly_invest"]},-{nm["cur_invest"]},C{NEED_TOTAL}-{nm["cur_cash"]})/12,0)&"歳","計算不能"))',
        "資産が貯まる年齢まで働く"); row += 1

    # C. FIRE後副業を増やす
    step5_row(ws, row,
        "C. FIRE後副業を増やす",
        f'=TEXT(ROUND({nm["fire_side"]}/10000,1),"0.0")&"万円/月"',
        f'=IF(C{ACTUAL_AT_TARGET}>=C{NEED_TOTAL},"現状で達成済",TEXT(ROUND(MAX({nm["fire_side"]},(C{FIRE_YEAR_LIVING}-C{ACTUAL_AT_TARGET}/25)/12)/10000,1),"0.0")&"万円/月")',
        "セミリタイア後の月収"); row += 1

    # D. FIRE後生活費を下げる
    step5_row(ws, row,
        "D. FIRE後生活費を下げる",
        f'=TEXT(ROUND({nm["fire_living"]}/10000,1),"0.0")&"万円/月"',
        f'=IF(C{ACTUAL_AT_TARGET}>=C{NEED_TOTAL},"現状で達成済",TEXT(ROUND(MIN({nm["fire_living"]},(C{ACTUAL_AT_TARGET}/25+C{FIRE_YEAR_SIDE})/12)/10000,1),"0.0")&"万円/月")',
        "『生活費』シートのFIRE後を調整"); row += 1

    # =================== 年次推移 ===================
    ws = wb.create_sheet("年次推移")
    ws.sheet_properties.tabColor = "BBDEFB"
    for letter, w in zip("ABCDEFGHI", [4, 8, 8, 14, 16, 16, 16, 18, 4]):
        ws.column_dimensions[letter].width = w
    title_row(ws, "📊 年次推移｜年齢別の資産", "現状ペース継続したら各年齢でいくら？")
    row = 4
    section_bar(ws, row, "現状ペース継続したら各年齢でいくら？", span=9); row += 1
    headers = ["年齢", "経過", "💧 貯金", "📈 運用FV", "💰 総資産", "総資産（万円表示）", "必要資産との差"]
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=2 + i, value=h)
        c.font = H1; c.fill = FILL_HEADER; c.alignment = CTR; c.border = B_THIN
    row += 1
    for n in range(0, 41):
        ws.cell(row=row, column=2, value=f"={nm['current_age']}+{n}").number_format = "0\"歳\""
        ws.cell(row=row, column=2).alignment = CTR; ws.cell(row=row, column=2).border = B_THIN
        ws.cell(row=row, column=3, value=n).number_format = "0\"年後\""
        ws.cell(row=row, column=3).alignment = CTR; ws.cell(row=row, column=3).border = B_THIN
        ws.cell(row=row, column=4, value=f"={nm['cur_cash']}").number_format = "#,##0\"円\""
        ws.cell(row=row, column=4).fill = FILL_SAVINGS
        ws.cell(row=row, column=4).alignment = RGT; ws.cell(row=row, column=4).border = B_THIN
        if n == 0:
            ws.cell(row=row, column=5, value=f"={nm['cur_invest']}")
        else:
            ws.cell(row=row, column=5, value=f"=FV({nm['rate']}/12,{n}*12,-{nm['monthly_invest']},-{nm['cur_invest']})")
        ws.cell(row=row, column=5).number_format = "#,##0\"円\""
        ws.cell(row=row, column=5).fill = FILL_INVEST
        ws.cell(row=row, column=5).alignment = RGT; ws.cell(row=row, column=5).border = B_THIN
        ws.cell(row=row, column=6, value=f"=D{row}+E{row}").number_format = "#,##0\"円\""
        ws.cell(row=row, column=6).font = LABEL
        ws.cell(row=row, column=6).alignment = RGT; ws.cell(row=row, column=6).border = B_THIN
        # 万円表示
        mc = ws.cell(row=row, column=7, value=man_formula(f"F{row}"))
        mc.font = MAN; mc.alignment = RGT; mc.fill = FILL_MAN; mc.border = B_THIN
        ws.cell(row=row, column=8, value=f"=F{row}-結論!C{NEED_TOTAL}")
        ws.cell(row=row, column=8).number_format = "#,##0\"円\";[Red]-#,##0\"円\""
        ws.cell(row=row, column=8).alignment = RGT; ws.cell(row=row, column=8).border = B_THIN
        row += 1

    # =================== FIRE試算 ===================
    ws = wb.create_sheet("FIRE試算")
    ws.sheet_properties.tabColor = "D1C4E9"
    for letter, w in zip("ABCDEFGH", [4, 22, 16, 12, 14, 18, 18, 30]):
        ws.column_dimensions[letter].width = w
    title_row(ws, "🔄 FIRE試算｜5パターン比較")
    row = 4
    section_bar(ws, row, "月積立や期間を変えるとどうなる？", span=8); row += 1
    headers = ["シナリオ", "月積立", "年利", "期間(年)", "予想総資産", "予想総資産（万円）", "備考"]
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=2 + i, value=h)
        c.font = H1; c.fill = FILL_HEADER; c.alignment = CTR; c.border = B_THIN
    row += 1
    scenarios = [
        ("①現状ペース", f"={nm['monthly_invest']}", f"={nm['rate']}", 10, "あなたの入力値そのまま"),
        ("②積立 +5万", f"={nm['monthly_invest']}+50000", f"={nm['rate']}", 10, "月積立を5万増やす"),
        ("③年利 +1%", f"={nm['monthly_invest']}", f"={nm['rate']}+0.01", 10, "投資配分をやや攻めに"),
        ("④期間 +5年", f"={nm['monthly_invest']}", f"={nm['rate']}", 15, "5年長く働く"),
        ("⑤積立0で運用のみ", 0, f"={nm['rate']}", 10, "今ある資産だけ運用"),
    ]
    for label, invest, rate, years, memo in scenarios:
        ws.cell(row=row, column=2, value=label).font = LABEL
        ws.cell(row=row, column=2).border = B_THIN
        ws.cell(row=row, column=3, value=invest).number_format = "#,##0\"円\""
        ws.cell(row=row, column=3).alignment = RGT; ws.cell(row=row, column=3).border = B_THIN
        ws.cell(row=row, column=4, value=rate).number_format = "0.0%"
        ws.cell(row=row, column=4).alignment = CTR; ws.cell(row=row, column=4).border = B_THIN
        ws.cell(row=row, column=5, value=years).number_format = "0\"年\""
        ws.cell(row=row, column=5).alignment = CTR; ws.cell(row=row, column=5).border = B_THIN
        result = ws.cell(row=row, column=6, value=f"=FV(D{row}/12,E{row}*12,-C{row},-{nm['cur_invest']})+{nm['cur_cash']}")
        result.fill = FILL_OUTPUT; result.border = B_BOLD; result.font = LABEL
        result.alignment = RGT; result.number_format = "#,##0\"円\""
        # 万円表示
        mc = ws.cell(row=row, column=7, value=man_formula(f"F{row}"))
        mc.font = MAN_BIG; mc.alignment = RGT; mc.fill = FILL_OUTPUT; mc.border = B_BOLD
        ws.cell(row=row, column=8, value=memo).font = SMALL
        ws.cell(row=row, column=8).alignment = LFT; ws.cell(row=row, column=8).border = B_THIN
        row += 1

    # =================== らくだ先生について ===================
    ws = wb.create_sheet("らくだ先生について")
    ws.sheet_properties.tabColor = BRAND_BEIGE
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 95
    info = [
        ("🐪", TITLE, "残業嫌いのらくだ先生"),
        ("", NORMAL, ""),
        ("■", H2, "プロフィール"),
        ("", NORMAL, "公立中学校教員（国語・20代）"),
        ("", NORMAL, "ICT・AIで残業ゼロ ｜ 副業10万 ｜ 5年で2000万円貯金"),
        ("", NORMAL, "日本株デイトレで-50万やった失敗から、低リスクインデックス投資に振り切り。"),
        ("", NORMAL, ""),
        ("■", H2, "発信中"),
        ("", NORMAL, "X (Twitter): @rakuda_sensei"),
        ("", NORMAL, "note: https://note.com/rakuda_sensei"),
        ("", NORMAL, "BOOTH: https://rakuda-sensei.booth.pm"),
        ("", NORMAL, ""),
        ("■", H2, "ご利用について"),
        ("", NORMAL, "・本商品の個人利用は自由"),
        ("", NORMAL, "・二次配布・転売は禁止"),
        ("", NORMAL, "・本商品の数式・テンプレで生じた結果について作者は責任を負いません"),
        ("", NORMAL, "・改善要望・不具合報告は X DM までお気軽に"),
    ]
    for i, (icon, font, text) in enumerate(info, start=1):
        ws.cell(row=i, column=1, value=icon)
        c = ws.cell(row=i, column=2, value=text)
        c.font = font; c.alignment = LFT

    out = OUT_DIR / "夫婦で実践｜サイドFIRE計画シート.xlsx"
    wb.save(out)
    return out


if __name__ == "__main__":
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out = build()
    print(f"✅ 完成: {out.name} ({out.stat().st_size // 1024}KB)")
