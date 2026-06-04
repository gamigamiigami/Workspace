#!/usr/bin/env python3
"""
ユーザー商品3つを「見やすさ・分かりやすさ」最優先で完全リビルド

設計原則:
1. 縦方向セクション分け (入力 → 自動計算 → アウトプット)
2. 入力セル=クリーム背景、計算セル=グレー背景で視覚区別
3. 単位を必ず明示 (円・万円・%・歳)
4. ヘッダー: ブランド茶色背景・白文字
5. 数式の意味をコメント列で説明
6. メイリオフォント統一・列幅統一
"""

from __future__ import annotations

import shutil
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

ROOT = Path(__file__).resolve().parents[3]
OUT_DIR = ROOT / "projects" / "rakuda-sensei" / "products" / "digital"

# ===== ブランドカラー =====
BRAND_BROWN = "8B6F47"     # ヘッダー濃茶
BRAND_BEIGE = "E8D5A7"     # セクションタイトル
BRAND_CREAM = "FDF6E3"     # 入力欄背景
BRAND_DARK = "3D2F1F"      # 濃茶文字
LIGHT_GRAY = "F0F0F0"      # 計算結果背景
SUCCESS_GREEN = "C8E6C9"   # アウトプット背景

# ===== Style 部品 =====
def F(size=11, bold=False, color=BRAND_DARK, italic=False):
    return Font(name="メイリオ", size=size, bold=bold, color=color, italic=italic)

TITLE_FONT = F(size=18, bold=True)
HEADER_FONT = F(size=11, bold=True, color="FFFFFF")
SECTION_FONT = F(size=13, bold=True, color=BRAND_BROWN)
LABEL_FONT = F(size=11, bold=True)
NORMAL_FONT = F(size=11)
NOTE_FONT = F(size=10, italic=True, color="666666")

HEADER_FILL = PatternFill(start_color=BRAND_BROWN, end_color=BRAND_BROWN, fill_type="solid")
SECTION_FILL = PatternFill(start_color=BRAND_BEIGE, end_color=BRAND_BEIGE, fill_type="solid")
INPUT_FILL = PatternFill(start_color=BRAND_CREAM, end_color=BRAND_CREAM, fill_type="solid")
CALC_FILL = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
OUTPUT_FILL = PatternFill(start_color=SUCCESS_GREEN, end_color=SUCCESS_GREEN, fill_type="solid")

THIN = Side(border_style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
THICK = Side(border_style="medium", color=BRAND_BROWN)
THICK_BORDER = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def section_header(ws, row: int, title: str, span_cols: int = 5):
    """セクション見出しを描画"""
    ws.cell(row=row, column=1, value=title).font = SECTION_FONT
    for c in range(1, span_cols + 1):
        ws.cell(row=row, column=c).fill = SECTION_FILL
        ws.cell(row=row, column=c).border = BORDER
    ws.row_dimensions[row].height = 24


def input_cell(ws, row, col, value, fmt="#,##0", note=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = INPUT_FILL
    c.border = THICK_BORDER
    c.font = LABEL_FONT
    c.alignment = RIGHT
    c.number_format = fmt
    if note:
        ws.cell(row=row, column=col + 1, value=note).font = NOTE_FONT
    return c


def calc_cell(ws, row, col, formula, fmt="#,##0"):
    c = ws.cell(row=row, column=col, value=formula)
    c.fill = CALC_FILL
    c.border = BORDER
    c.font = NORMAL_FONT
    c.alignment = RIGHT
    c.number_format = fmt
    return c


def output_cell(ws, row, col, formula, fmt="#,##0"):
    c = ws.cell(row=row, column=col, value=formula)
    c.fill = OUTPUT_FILL
    c.border = THICK_BORDER
    c.font = F(size=13, bold=True)
    c.alignment = RIGHT
    c.number_format = fmt
    return c


def label_cell(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = LABEL_FONT
    c.alignment = LEFT
    c.border = BORDER
    return c


def add_title(ws, title: str, subtitle: str = ""):
    ws.cell(row=1, column=1, value="🐪").font = F(size=24)
    c = ws.cell(row=1, column=2, value=title)
    c.font = TITLE_FONT
    ws.row_dimensions[1].height = 32
    if subtitle:
        s = ws.cell(row=2, column=2, value=subtitle)
        s.font = NOTE_FONT


def add_brand_footer_sheet(wb: Workbook):
    if "らくだ先生について" in wb.sheetnames:
        del wb["らくだ先生について"]
    ws = wb.create_sheet("らくだ先生について")
    ws.sheet_properties.tabColor = BRAND_BEIGE
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 95
    info = [
        ("🐪", TITLE_FONT, "残業嫌いのらくだ先生"),
        ("", NORMAL_FONT, ""),
        ("■", SECTION_FONT, "プロフィール"),
        ("", NORMAL_FONT, "公立中学校教員（国語・20代）"),
        ("", NORMAL_FONT, "ICT・AIで残業ゼロ ｜ 副業10万 ｜ 5年で2000万円貯金"),
        ("", NORMAL_FONT, "日本株デイトレで-50万やった失敗から、低リスクインデックス投資に振り切り。"),
        ("", NORMAL_FONT, ""),
        ("■", SECTION_FONT, "発信中"),
        ("", NORMAL_FONT, "X (Twitter): @rakuda_sensei"),
        ("", NORMAL_FONT, "note: https://note.com/rakuda_sensei"),
        ("", NORMAL_FONT, "BOOTH: https://rakuda-sensei.booth.pm"),
        ("", NORMAL_FONT, ""),
        ("■", SECTION_FONT, "ご利用について"),
        ("", NORMAL_FONT, "・本商品の個人利用・校内利用は自由"),
        ("", NORMAL_FONT, "・二次配布・転売は禁止"),
        ("", NORMAL_FONT, "・本商品の数式・テンプレで生じた結果について作者は責任を負いません"),
        ("", NORMAL_FONT, "・改善要望・不具合報告は X DM までお気軽に！"),
    ]
    for i, (icon, font, text) in enumerate(info, start=1):
        ws.cell(row=i, column=1, value=icon)
        c = ws.cell(row=i, column=2, value=text)
        c.font = font
        c.alignment = LEFT


# =========================================================================
# 1. 伊神流FP (リビルド版)
# =========================================================================
def rebuild_fp() -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    # ========== 使い方 ==========
    ws = wb.create_sheet("使い方", 0)
    ws.sheet_properties.tabColor = BRAND_BROWN
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 95
    add_title(ws, "伊神流FP｜家計と先取り貯金テンプレ",
              "「先取り貯金 → 残ったお金で生活」の流れを Excel で見える化")

    lines = [
        "",
        "■ 使い方（5ステップ）",
        "1. シート「一人暮らし用」または「ルームシェア用」を開く",
        "2. 🟡黄色セル＝入力欄。手取り給料・ボーナス・固定費・生活費を埋める",
        "3. ⬜灰色セル＝自動計算（触らない）",
        "4. 🟩緑セル＝最終アウトプット（先取り貯金額・実質貯金額）",
        "5. 目標貯金額（1〜5）を選んで、自分のペースを決める",
        "",
        "■ シート構成",
        "・一人暮らし用: 単身者向けの計算テンプレ",
        "・ルームシェア用: 同居人と生活費を分担するパターン",
        "・用語集: 数式に出てくる用語の説明",
        "",
        "■ 実際の運用",
        "・給料が入ったらまず「先取り貯金額」を貯金口座に振込",
        "・「目標小遣い」と「小遣い合計」を見比べて使い方を決める",
        "・残った小遣いを有意義に使い、月末に残ったらまた貯金口座へ",
        "",
        "■ 注意",
        "→ 入った数字はサンプルです。あなたの状況に書き換えてください",
        "→ NISA・iDeCo は元本割れリスクあり、必ず自己責任で判断",
    ]
    for i, line in enumerate(lines, start=3):
        c = ws.cell(row=i, column=2, value=line)
        if line.startswith("■"):
            c.font = SECTION_FONT
        elif line.startswith("→"):
            c.font = NOTE_FONT
        else:
            c.font = NORMAL_FONT
        c.alignment = LEFT

    # ========== 一人暮らし用 ==========
    def build_household(sheet_name: str, tab_color: str, default_income: int, default_bonus: int, has_share: bool = False):
        ws = wb.create_sheet(sheet_name)
        ws.sheet_properties.tabColor = tab_color
        for col_letter, width in zip("ABCDE", [22, 18, 24, 18, 32]):
            ws.column_dimensions[col_letter].width = width

        add_title(ws, sheet_name, "🟡入力 → ⬜自動計算 → 🟩アウトプット の順に見ます")

        row = 4

        # === セクション1: 収入 ===
        section_header(ws, row, "① 収入を入力（🟡黄色セルだけ書き換え）"); row += 1
        label_cell(ws, row, 1, "月手取り給料"); input_cell(ws, row, 2, default_income, "#,##0\"円\"", "(税金・社保引いた額)"); row += 1
        label_cell(ws, row, 1, "年間ボーナス（手取り）"); input_cell(ws, row, 2, default_bonus, "#,##0\"円\"", "(年2回合計)"); row += 1
        if has_share:
            label_cell(ws, row, 1, "シェア人数"); input_cell(ws, row, 2, 2, "0\"人\"", "(自分含む)"); row += 1
        row += 1

        # === セクション2: 固定費 ===
        section_header(ws, row, "② 固定費を入力（毎月絶対かかる費用）"); row += 1
        fixed_start = row
        ws.cell(row=row, column=1, value="項目").font = HEADER_FONT
        ws.cell(row=row, column=1).fill = HEADER_FILL
        ws.cell(row=row, column=2, value="月額").font = HEADER_FONT
        ws.cell(row=row, column=2).fill = HEADER_FILL
        row += 1
        for item, amount in [("携帯代", 4000), ("保険料", 3000), ("ガソリン代", 6000), ("サブスク", 3000), ("奨学金", 9000)]:
            label_cell(ws, row, 1, item); input_cell(ws, row, 2, amount, "#,##0\"円\"")
            row += 1
        # 入力可空き行
        for _ in range(2):
            label_cell(ws, row, 1, ""); input_cell(ws, row, 2, "", "#,##0\"円\"")
            row += 1
        fixed_end = row - 1
        label_cell(ws, row, 1, "🔢 固定費 合計"); calc_cell(ws, row, 2, f"=SUM(B{fixed_start+1}:B{fixed_end})", "#,##0\"円\"")
        fixed_total_row = row
        row += 2

        # === セクション3: 生活費 ===
        section_header(ws, row, "③ 生活費を入力（住居系の費用）"); row += 1
        living_start = row
        ws.cell(row=row, column=1, value="項目").font = HEADER_FONT
        ws.cell(row=row, column=1).fill = HEADER_FILL
        ws.cell(row=row, column=2, value="月額").font = HEADER_FONT
        ws.cell(row=row, column=2).fill = HEADER_FILL
        row += 1
        living_items = [("家賃", 76500), ("共益費", 2200), ("駐車場", 8800), ("通信費", 6000), ("水道", 3000),
                        ("ガス", 2000), ("電気", 5000), ("食費", 30000), ("家賃手当 (引)", -27000)]
        for item, amount in living_items:
            label_cell(ws, row, 1, item); input_cell(ws, row, 2, amount, "#,##0\"円\"")
            row += 1
        living_end = row - 1
        if has_share:
            label_cell(ws, row, 1, "🔢 生活費 合計（世帯）"); calc_cell(ws, row, 2, f"=SUM(B{living_start+1}:B{living_end})", "#,##0\"円\"")
            row += 1
            label_cell(ws, row, 1, "🔢 生活費 一人あたり"); calc_cell(ws, row, 2, f"=B{row-1}/B6", "#,##0\"円\"")
            living_share_row = row
            row += 1
            living_total_per_person_row = living_share_row
        else:
            label_cell(ws, row, 1, "🔢 生活費 合計"); calc_cell(ws, row, 2, f"=SUM(B{living_start+1}:B{living_end})", "#,##0\"円\"")
            living_total_per_person_row = row
        row += 2

        # === セクション4: 計算結果 (自動) ===
        section_header(ws, row, "④ 自動計算（⬜灰色セル）"); row += 1
        income_row = 5  # 月手取り
        bonus_row = 6   # 年間ボーナス
        label_cell(ws, row, 1, "毎月の支出合計（固定費＋生活費）"); calc_cell(ws, row, 2, f"=B{fixed_total_row}+B{living_total_per_person_row}", "#,##0\"円\""); month_expense_row = row; row += 1
        label_cell(ws, row, 1, "毎月の貯金可能額"); calc_cell(ws, row, 2, f"=B{income_row}-B{month_expense_row}", "#,##0\"円\""); month_savable_row = row; row += 1
        label_cell(ws, row, 1, "年間の貯金可能額（ボーナス込）"); calc_cell(ws, row, 2, f"=B{month_savable_row}*12+B{bonus_row}", "#,##0\"円\""); year_savable_row = row; row += 2

        # === セクション5: 目標貯金額を選ぶ ===
        section_header(ws, row, "⑤ 目標貯金総額を選ぶ（1〜5）"); row += 1
        ws.cell(row=row, column=1, value="選択").font = HEADER_FONT
        ws.cell(row=row, column=1).fill = HEADER_FILL
        ws.cell(row=row, column=2, value="目標額").font = HEADER_FONT
        ws.cell(row=row, column=2).fill = HEADER_FILL
        ws.cell(row=row, column=3, value="月必要貯金").font = HEADER_FONT
        ws.cell(row=row, column=3).fill = HEADER_FILL
        ws.cell(row=row, column=4, value="月の小遣い目標").font = HEADER_FONT
        ws.cell(row=row, column=4).fill = HEADER_FILL
        ws.cell(row=row, column=5, value="達成までの月数").font = HEADER_FONT
        ws.cell(row=row, column=5).fill = HEADER_FILL
        row += 1
        target_amounts = [1000000, 1500000, 1800000, 2000000, 2500000]
        target_start = row
        for n, amount in enumerate(target_amounts, start=1):
            label_cell(ws, row, 1, f"{n}")
            input_cell(ws, row, 2, amount, "#,##0\"円\"")
            calc_cell(ws, row, 3, f"=ROUND((B{row}-B{bonus_row})/12,0)", "#,##0\"円\"")
            calc_cell(ws, row, 4, f"=ROUND((B{income_row}-B{month_expense_row}-C{row})/1000,0)*1000", "#,##0\"円\"")
            calc_cell(ws, row, 5, f"=ROUND(B{row}/B{month_savable_row},1)", "0.0\"ヶ月\"")
            row += 1
        target_end = row - 1
        row += 1

        label_cell(ws, row, 1, "↓どの目標を採用？ 番号を入力"); input_cell(ws, row, 2, 4, "0", "(1〜5から選択)"); chosen_target_row = row; row += 1
        label_cell(ws, row, 1, "  → 採用された目標小遣い"); calc_cell(ws, row, 2, f"=INDEX(D{target_start}:D{target_end},B{chosen_target_row})", "#,##0\"円\""); target_okozukai_row = row; row += 2

        # === セクション6: 投資配分 ===
        section_header(ws, row, "⑥ 投資配分（NISA / iDeCo を埋める）"); row += 1
        label_cell(ws, row, 1, "NISA 月積立"); input_cell(ws, row, 2, 30000, "#,##0\"円\""); nisa_row = row; row += 1
        label_cell(ws, row, 1, "iDeCo 月積立"); input_cell(ws, row, 2, 0, "#,##0\"円\""); ideco_row = row; row += 2

        # === セクション7: アウトプット (緑) ===
        section_header(ws, row, "⑦ アウトプット（🟩緑セル）"); row += 1
        label_cell(ws, row, 1, "💰 先取り貯金額（給料日に即振込）")
        output_cell(ws, row, 2, f"=ROUNDDOWN((B{income_row}-B{month_expense_row}-B{target_okozukai_row})/1000,0)*1000-B{nisa_row}-B{ideco_row}", "#,##0\"円\"")
        row += 1
        label_cell(ws, row, 1, "💰 投資込みの月貯蓄合計")
        output_cell(ws, row, 2, f"=B{row-1}+B{nisa_row}+B{ideco_row}", "#,##0\"円\"")
        row += 1
        label_cell(ws, row, 1, "💰 実質年間貯蓄額（ボーナス込）")
        output_cell(ws, row, 2, f"=B{row}*12+B{bonus_row}", "#,##0\"円\"")
        row += 1
        label_cell(ws, row, 1, "💰 5年で貯まる累計（年利0%換算）")
        output_cell(ws, row, 2, f"=B{row-1}*5", "#,##0\"円\"")

        return ws

    build_household("一人暮らし用", "FFCDD2", 200000, 1000000, has_share=False)
    build_household("ルームシェア用", "BBDEFB", 190000, 0, has_share=True)

    # ========== 用語集 ==========
    ws = wb.create_sheet("用語集")
    ws.sheet_properties.tabColor = BRAND_BEIGE
    for col_letter, width in zip("ABC", [4, 22, 70]):
        ws.column_dimensions[col_letter].width = width
    add_title(ws, "用語集")
    row = 4
    for term, desc in [
        ("固定費", "携帯・保険・サブスクなど、毎月絶対かかる費用"),
        ("生活費", "家賃・光熱費・食費など、住居系の費用"),
        ("貯金可能額", "手取り − 固定費 − 生活費"),
        ("必要貯金額(月)", "目標貯金総額を達成するために、毎月貯めるべき金額"),
        ("目標小遣い", "手取り − 固定費 − 生活費 − 必要貯金額"),
        ("実質貯金額", "目標小遣い × 12 + ボーナス（年間トータル貯蓄）"),
        ("先取り貯金額", "給料が入ったら真っ先に貯金口座に振り込む金額"),
        ("NISA", "投資信託の利益が非課税になる制度（年360万まで）"),
        ("iDeCo", "個人型確定拠出年金。掛金が全額所得控除になる制度"),
    ]:
        c = ws.cell(row=row, column=2, value=term); c.font = LABEL_FONT; c.fill = SECTION_FILL; c.border = BORDER
        c = ws.cell(row=row, column=3, value=desc); c.font = NORMAL_FONT; c.alignment = LEFT; c.border = BORDER
        row += 1

    add_brand_footer_sheet(wb)
    out = OUT_DIR / "伊神流FP｜家計と先取り貯金テンプレ.xlsx"
    wb.save(out)
    return out


# =========================================================================
# 2. サイドFIRE計画 (リビルド版)
# =========================================================================
def rebuild_fire() -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    # ========== 使い方 ==========
    ws = wb.create_sheet("使い方", 0)
    ws.sheet_properties.tabColor = BRAND_BROWN
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 95
    add_title(ws, "夫婦で実践｜サイドFIRE計画シート",
              "夫婦（または2人）で「いつ・いくらでセミリタイア」を試算")

    lines = [
        "",
        "■ サイドFIREとは",
        "資産収入だけでは生活費を全部賄えない分を、緩めの労働で補う戦略。",
        "完全FIRE より早く・現実的に到達できる。バリスタFIREとも呼ばれる。",
        "",
        "■ シート構成",
        "・初期設定: 現在年齢・設定年・ライフイベント",
        "・自分(一人目): あなたの月手取り・生活費・現在資産・投資設定",
        "・パートナー(二人目): もう一人の同上",
        "・合算: 夫婦の合計と「いつFIREできるか」の年齢別シミュレーション",
        "・資産収入: 資産額別の年利・月利の利益試算",
        "・生活費: 月の生活費内訳（家計簿）",
        "",
        "■ 使い方（4ステップ）",
        "1. 「初期設定」で現在年齢・設定年を入力",
        "2. 「自分」「パートナー」で月手取り・月生活費・現在資産を入力",
        "3. 投資設定（年利・初期投資・月積立）を入力",
        "4. 「合算」で何歳でFIRE達成できるか確認",
        "",
        "■ 4つのFIRE試算ロジック",
        "・FIRE① 積立有 : 月積立を続けた場合の○年後の資産",
        "・FIRE② 積立0 : 積立をストップしてからの○年後の資産（運用のみ）",
        "・FIRE③ 取崩し: 月◯円取崩しを始めて○年後に残る資産",
        "・FIRE④ 寿命  : 資産が尽きるまで何年もつか（NPER関数）",
        "",
        "■ 注意",
        "→ 入った数字はサンプルです。自分の状況に書き換えてください",
        "→ 想定年利は過去実績ベースの参考値、投資は元本割れリスクあり",
        "→ 年金額は2026年時点の概算、制度改正で変動します",
    ]
    for i, line in enumerate(lines, start=3):
        c = ws.cell(row=i, column=2, value=line)
        if line.startswith("■"):
            c.font = SECTION_FONT
        elif line.startswith("→"):
            c.font = NOTE_FONT
        else:
            c.font = NORMAL_FONT
        c.alignment = LEFT

    # ========== 初期設定 ==========
    ws = wb.create_sheet("初期設定")
    ws.sheet_properties.tabColor = BRAND_BEIGE
    for letter, w in zip("ABCD", [22, 16, 6, 50]):
        ws.column_dimensions[letter].width = w
    add_title(ws, "初期設定")
    row = 4
    section_header(ws, row, "あなたの基本情報（🟡黄色セル）", span_cols=4); row += 1
    label_cell(ws, row, 1, "現在年齢"); input_cell(ws, row, 2, 27, "0\"歳\"", "(あなたの年齢)"); age_row = row; row += 1
    label_cell(ws, row, 1, "設定年"); input_cell(ws, row, 2, 2026, "0\"年\"", "(計算の起点となる年)"); row += 1
    row += 1
    section_header(ws, row, "ライフイベント（任意）", span_cols=4); row += 1
    label_cell(ws, row, 1, "出産予定年"); input_cell(ws, row, 2, 2029, "0\"年\""); row += 1
    label_cell(ws, row, 1, "住宅購入予定年"); input_cell(ws, row, 2, "", "0\"年\""); row += 1
    label_cell(ws, row, 1, "子供独立予定年"); input_cell(ws, row, 2, "", "0\"年\"")

    # ========== 自分 / パートナー ==========
    def build_person_sheet(sheet_name: str, tab_color: str, default_income: int, default_bonus: int,
                           default_living: int, default_assets: int, default_rate: float):
        ws = wb.create_sheet(sheet_name)
        ws.sheet_properties.tabColor = tab_color
        for letter, w in zip("ABCD", [24, 16, 6, 42]):
            ws.column_dimensions[letter].width = w
        add_title(ws, sheet_name)
        row = 4

        # 収入
        section_header(ws, row, "① 収入（月単位）", span_cols=4); row += 1
        label_cell(ws, row, 1, "月手取り給料"); input_cell(ws, row, 2, default_income, "#,##0\"円\""); month_income_row = row; row += 1
        label_cell(ws, row, 1, "ボーナス（年合計手取り）"); input_cell(ws, row, 2, default_bonus, "#,##0\"円\""); bonus_row = row; row += 1
        label_cell(ws, row, 1, "🔢 年手取り（自動）"); calc_cell(ws, row, 2, f"=B{month_income_row}*12+B{bonus_row}", "#,##0\"円\""); year_income_row = row; row += 2

        # 支出
        section_header(ws, row, "② 月の生活費", span_cols=4); row += 1
        label_cell(ws, row, 1, "月生活費合計"); input_cell(ws, row, 2, default_living, "#,##0\"円\""); month_living_row = row; row += 1
        label_cell(ws, row, 1, "🔢 月貯蓄"); calc_cell(ws, row, 2, f"=B{month_income_row}-B{month_living_row}", "#,##0\"円\""); month_savings_row = row; row += 1
        label_cell(ws, row, 1, "🔢 年貯蓄（ボーナス込）"); calc_cell(ws, row, 2, f"=B{month_savings_row}*12+B{bonus_row}", "#,##0\"円\""); year_savings_row = row; row += 2

        # 資産
        section_header(ws, row, "③ 現在の資産", span_cols=4); row += 1
        label_cell(ws, row, 1, "現在の貯金額"); input_cell(ws, row, 2, default_assets, "#,##0\"円\""); current_savings_row = row; row += 1
        label_cell(ws, row, 1, "セミリタイアまでの年数"); input_cell(ws, row, 2, 10, "0\"年\""); years_row = row; row += 1
        label_cell(ws, row, 1, "🔢 セミリタイア時点の貯金"); calc_cell(ws, row, 2, f"=B{current_savings_row}+B{year_savings_row}*B{years_row}", "#,##0\"円\""); future_savings_row = row; row += 2

        # 投資
        section_header(ws, row, "④ 投資（運用）設定", span_cols=4); row += 1
        label_cell(ws, row, 1, "初期投資元本"); input_cell(ws, row, 2, default_assets, "#,##0\"円\""); init_invest_row = row; row += 1
        label_cell(ws, row, 1, "月積立額"); input_cell(ws, row, 2, 50000, "#,##0\"円\""); month_invest_row = row; row += 1
        label_cell(ws, row, 1, "想定年利"); input_cell(ws, row, 2, default_rate, "0.0%"); rate_row = row; row += 1
        label_cell(ws, row, 1, "運用期間"); calc_cell(ws, row, 2, f"=B{years_row}", "0\"年\""); period_row = row; row += 1
        label_cell(ws, row, 1, "🔢 セミリタイア時点の運用資産"); output_cell(ws, row, 2, f"=FV(B{rate_row}/12,B{period_row}*12,-B{month_invest_row},-B{init_invest_row})", "#,##0\"円\""); invest_fv_row = row; row += 1
        label_cell(ws, row, 1, "🔢 総資産（貯金＋運用）"); output_cell(ws, row, 2, f"=B{future_savings_row}+B{invest_fv_row}", "#,##0\"円\"")

        return {
            "sheet": sheet_name,
            "month_income": month_income_row,
            "year_income": year_income_row,
            "year_savings": year_savings_row,
            "current_savings": current_savings_row,
            "years": years_row,
            "future_savings": future_savings_row,
            "month_invest": month_invest_row,
            "rate": rate_row,
            "period": period_row,
            "invest_fv": invest_fv_row,
            "init_invest": init_invest_row,
        }

    person_a = build_person_sheet("自分", "C8E6C9", 250000, 1000000, 100000, 1500000, 0.05)
    person_b = build_person_sheet("パートナー", "BBDEFB", 200000, 600000, 80000, 500000, 0.04)

    # ========== 合算 ==========
    ws = wb.create_sheet("合算")
    ws.sheet_properties.tabColor = "FFE0B2"
    for letter, w in zip("ABCDEF", [22, 18, 18, 18, 18, 32]):
        ws.column_dimensions[letter].width = w
    add_title(ws, "合算｜夫婦のFIREシミュレーション")

    row = 4
    section_header(ws, row, "① 現状の合計", span_cols=6); row += 1
    label_cell(ws, row, 1, "夫婦の年間貯蓄合計")
    calc_cell(ws, row, 2, f"=自分!B{person_a['year_savings']}+パートナー!B{person_b['year_savings']}", "#,##0\"円\"")
    couple_year_savings = row
    row += 1
    label_cell(ws, row, 1, "夫婦の現在資産（運用元本）")
    calc_cell(ws, row, 2, f"=自分!B{person_a['init_invest']}+パートナー!B{person_b['init_invest']}", "#,##0\"円\"")
    couple_init = row
    row += 1
    label_cell(ws, row, 1, "夫婦の月積立合計")
    calc_cell(ws, row, 2, f"=自分!B{person_a['month_invest']}+パートナー!B{person_b['month_invest']}", "#,##0\"円\"")
    couple_invest = row
    row += 1
    label_cell(ws, row, 1, "夫婦の平均年利")
    calc_cell(ws, row, 2, f"=AVERAGE(自分!B{person_a['rate']},パートナー!B{person_b['rate']})", "0.0%")
    couple_rate = row
    row += 2

    section_header(ws, row, "② FIRE 4シナリオ（年数を変えて試算）", span_cols=6); row += 1
    ws.cell(row=row, column=1, value="シナリオ").font = HEADER_FONT; ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=2, value="月積立").font = HEADER_FONT; ws.cell(row=row, column=2).fill = HEADER_FILL
    ws.cell(row=row, column=3, value="年利").font = HEADER_FONT; ws.cell(row=row, column=3).fill = HEADER_FILL
    ws.cell(row=row, column=4, value="期間（年）").font = HEADER_FONT; ws.cell(row=row, column=4).fill = HEADER_FILL
    ws.cell(row=row, column=5, value="資産（円）").font = HEADER_FONT; ws.cell(row=row, column=5).fill = HEADER_FILL
    ws.cell(row=row, column=6, value="メモ").font = HEADER_FONT; ws.cell(row=row, column=6).fill = HEADER_FILL
    row += 1
    # FIRE①積立有
    label_cell(ws, row, 1, "FIRE① 積立有")
    input_cell(ws, row, 2, f"=B{couple_invest}", "#,##0"); input_cell(ws, row, 3, f"=B{couple_rate}", "0.0%"); input_cell(ws, row, 4, 5, "0")
    output_cell(ws, row, 5, f"=FV(C{row}/12,D{row}*12,-B{row},-B{couple_init})", "#,##0\"円\"")
    label_cell(ws, row, 6, "5年継続したらの資産")
    fire1_row = row; row += 1
    # FIRE②積立0
    label_cell(ws, row, 1, "FIRE② 積立0")
    input_cell(ws, row, 2, 0, "#,##0"); input_cell(ws, row, 3, f"=B{couple_rate}", "0.0%"); input_cell(ws, row, 4, 5, "0")
    output_cell(ws, row, 5, f"=FV(C{row}/12,D{row}*12,0,-E{fire1_row})", "#,##0\"円\"")
    label_cell(ws, row, 6, "積立止めて運用のみ")
    fire2_row = row; row += 1
    # FIRE③取崩し
    label_cell(ws, row, 1, "FIRE③ 月取崩し")
    input_cell(ws, row, 2, 100000, "#,##0"); input_cell(ws, row, 3, f"=B{couple_rate}/2", "0.0%"); input_cell(ws, row, 4, 10, "0")
    output_cell(ws, row, 5, f"=FV(C{row}/12,D{row}*12,B{row},-E{fire2_row})", "#,##0\"円\"")
    label_cell(ws, row, 6, "10年取崩し続けたらの残高")
    fire3_row = row; row += 1
    # FIRE④寿命
    label_cell(ws, row, 1, "FIRE④ 寿命（NPER）")
    input_cell(ws, row, 2, 200000, "#,##0"); input_cell(ws, row, 3, 0.03, "0.0%"); label_cell(ws, row, 4, "")
    output_cell(ws, row, 5, f"=IFERROR(ROUND(NPER(C{row}/12,-B{row},E{fire3_row})/12,1),\"資金不足\")", "0.0\"年\"")
    label_cell(ws, row, 6, "資産が尽きるまで何年")
    row += 2

    # ========== 資産収入 ==========
    ws = wb.create_sheet("資産収入")
    ws.sheet_properties.tabColor = "D1C4E9"
    for letter, w in zip("ABCDE", [18, 14, 14, 14, 30]):
        ws.column_dimensions[letter].width = w
    add_title(ws, "資産収入｜資産額別の年利・月利")
    row = 4
    section_header(ws, row, "想定年利を入力（🟡）", span_cols=5); row += 1
    label_cell(ws, row, 1, "想定年利A"); input_cell(ws, row, 2, 0.03, "0.0%", "(保守的)"); rate_a = row; row += 1
    label_cell(ws, row, 1, "想定年利B"); input_cell(ws, row, 2, 0.05, "0.0%", "(らくだ式)"); rate_b = row; row += 2

    section_header(ws, row, "資産額別シミュレーション", span_cols=5); row += 1
    for c, h in enumerate(["資産額", "年利A_年益", "年利A_月益", "年利B_年益", "年利B_月益"]):
        ws.cell(row=row, column=c + 1, value=h).font = HEADER_FONT
        ws.cell(row=row, column=c + 1).fill = HEADER_FILL
        ws.cell(row=row, column=c + 1).alignment = CENTER
    row += 1
    for amount in [10000000, 20000000, 30000000, 50000000, 80000000, 100000000]:
        input_cell(ws, row, 1, amount, "#,##0\"円\"")
        calc_cell(ws, row, 2, f"=A{row}*B{rate_a}", "#,##0\"円\"")
        calc_cell(ws, row, 3, f"=B{row}/12", "#,##0\"円\"")
        calc_cell(ws, row, 4, f"=A{row}*B{rate_b}", "#,##0\"円\"")
        calc_cell(ws, row, 5, f"=D{row}/12", "#,##0\"円\"")
        row += 1

    # ========== 生活費 ==========
    ws = wb.create_sheet("生活費")
    ws.sheet_properties.tabColor = "FFCCBC"
    for letter, w in zip("ABCD", [22, 16, 6, 40]):
        ws.column_dimensions[letter].width = w
    add_title(ws, "生活費｜月の家計内訳")
    row = 4
    section_header(ws, row, "項目別月額", span_cols=4); row += 1
    ws.cell(row=row, column=1, value="項目").font = HEADER_FONT; ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=2, value="月額").font = HEADER_FONT; ws.cell(row=row, column=2).fill = HEADER_FILL
    ws.cell(row=row, column=4, value="備考").font = HEADER_FONT; ws.cell(row=row, column=4).fill = HEADER_FILL
    row += 1
    items = [
        ("住居費", 90000, "家賃・住宅ローン・管理費"),
        ("食費", 80000, "自炊中心＋外食少し"),
        ("水道光熱費", 22000, "電気・ガス・水道"),
        ("通信費", 18000, "スマホ2台＋ネット"),
        ("日用品", 12000, "消耗品・雑貨"),
        ("交通費", 20000, "車・公共交通"),
        ("教育費", 30000, "公立想定・習い事含む"),
        ("医療・保険", 15000, "医療費＋保険"),
        ("娯楽・交際費", 25000, "レジャー・外食"),
        ("社会保険料", 40000, ""),
    ]
    sum_start = row
    for item, amount, note in items:
        label_cell(ws, row, 1, item); input_cell(ws, row, 2, amount, "#,##0\"円\"")
        c = ws.cell(row=row, column=4, value=note); c.font = NOTE_FONT; c.border = BORDER; c.alignment = LEFT
        row += 1
    sum_end = row - 1
    label_cell(ws, row, 1, "🔢 合計（月）"); output_cell(ws, row, 2, f"=SUM(B{sum_start}:B{sum_end})", "#,##0\"円\"")

    add_brand_footer_sheet(wb)
    out = OUT_DIR / "夫婦で実践｜サイドFIRE計画シート.xlsx"
    wb.save(out)
    return out


# =========================================================================
# 3. Todo＆時間割 (リビルド版)
# =========================================================================
def rebuild_todo() -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    # 使い方
    ws = wb.create_sheet("使い方", 0)
    ws.sheet_properties.tabColor = BRAND_BROWN
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 95
    add_title(ws, "中学校教員のためのToDo＆時間割テンプレ", "校務の見える化・印刷して教卓に貼って使う")
    lines = [
        "",
        "■ シート構成",
        "・時間割: 週の授業コマ表（1〜6時限 × 月〜金）",
        "・ToDo: チェック付き・優先度・期限管理（30件）",
        "",
        "■ 使い方",
        "1. 「時間割」シートで自分のクラス名（例: 2-1, 3-A）を埋める",
        "2. 「ToDo」シートで毎日5件を上限に書く（多いと未消化が増える）",
        "3. 完了したらチェックボックスを✓",
        "4. 週末に未完了を翌週へ繰り越し",
        "",
        "■ 印刷",
        "A4横で印刷 → 教卓に貼ると、同僚から「便利そう」と声かけられがち。",
        "",
        "■ 運用のコツ",
        "・優先度A: 今日中 / B: 今週中 / C: 来週以降 でフィルタ可能",
        "・〆切は必ず「日付」で（『今週中』では先送りされる）",
        "・備考は『誰のために』『どんな成果が必要か』を1行で",
    ]
    for i, line in enumerate(lines, start=3):
        c = ws.cell(row=i, column=2, value=line)
        if line.startswith("■"):
            c.font = SECTION_FONT
        elif line.startswith("→") or line.startswith("A4"):
            c.font = NOTE_FONT
        else:
            c.font = NORMAL_FONT
        c.alignment = LEFT

    # 時間割
    ws = wb.create_sheet("時間割")
    ws.sheet_properties.tabColor = "C8E6C9"
    for letter, w in zip("ABCDEFG", [8, 18, 18, 18, 18, 18, 8]):
        ws.column_dimensions[letter].width = w
    add_title(ws, "週間時間割")
    row = 4
    # ヘッダー
    headers = ["時限", "月", "火", "水", "木", "金"]
    for c, h in enumerate(headers):
        cell = ws.cell(row=row, column=c + 1, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    row += 1
    for i in range(1, 7):
        ws.cell(row=row, column=1, value=f"{i}時限").font = LABEL_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        ws.cell(row=row, column=1).alignment = CENTER
        ws.cell(row=row, column=1).border = BORDER
        for col in range(2, 7):
            c = ws.cell(row=row, column=col, value="")
            c.fill = INPUT_FILL
            c.border = BORDER
            c.alignment = CENTER
            c.font = NORMAL_FONT
        ws.row_dimensions[row].height = 32
        row += 1
    row += 1
    # 行事・備考欄
    section_header(ws, row, "今週の行事・連絡事項", span_cols=6); row += 1
    for _ in range(5):
        c = ws.cell(row=row, column=1, value="")
        c.fill = INPUT_FILL
        c.border = BORDER
        ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=6)
        row += 1

    # ToDo
    ws = wb.create_sheet("ToDo")
    ws.sheet_properties.tabColor = "FFCDD2"
    for letter, w in zip("ABCDEFG", [10, 8, 40, 14, 30, 8, 8]):
        ws.column_dimensions[letter].width = w
    add_title(ws, "校務 ToDo リスト")

    row = 4
    headers = ["状態", "優先度", "やること", "〆切", "備考", "", ""]
    for c, h in enumerate(headers[:5]):
        cell = ws.cell(row=row, column=c + 1, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    row += 1
    # 30件分のテンプレ
    for i in range(30):
        c = ws.cell(row=row, column=1, value="")
        c.fill = INPUT_FILL
        c.alignment = CENTER
        c.border = BORDER
        c = ws.cell(row=row, column=2, value="")
        c.fill = INPUT_FILL
        c.alignment = CENTER
        c.border = BORDER
        c = ws.cell(row=row, column=3, value="")
        c.fill = INPUT_FILL
        c.alignment = LEFT
        c.border = BORDER
        c = ws.cell(row=row, column=4, value="")
        c.fill = INPUT_FILL
        c.alignment = CENTER
        c.border = BORDER
        c = ws.cell(row=row, column=5, value="")
        c.fill = INPUT_FILL
        c.alignment = LEFT
        c.border = BORDER
        row += 1

    # 状態・優先度のドロップダウン
    dv1 = DataValidation(type="list", formula1='"☐,進行中,✓完了"', allow_blank=True)
    dv1.add(f"A5:A{row-1}")
    ws.add_data_validation(dv1)
    dv2 = DataValidation(type="list", formula1='"A:今日中,B:今週中,C:来週以降"', allow_blank=True)
    dv2.add(f"B5:B{row-1}")
    ws.add_data_validation(dv2)

    add_brand_footer_sheet(wb)
    out = OUT_DIR / "中学校教員のためのToDo＆時間割テンプレ.xlsx"
    wb.save(out)
    return out


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"📁 出力先: {OUT_DIR}\n")

    products = [
        ("伊神流FP (リビルド)", rebuild_fp),
        ("サイドFIRE計画 (リビルド)", rebuild_fire),
        ("Todo&時間割 (リビルド)", rebuild_todo),
    ]
    for name, fn in products:
        try:
            out = fn()
            print(f"  ✅ {name} → {out.name} ({out.stat().st_size // 1024}KB)")
        except Exception as e:
            print(f"  ❌ {name}: {e}")
            import traceback
            traceback.print_exc()

    print("\n🎁 リビルド完了 (見やすさ最優先)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
