#!/usr/bin/env python3
"""
note 記事003「先取り投資→生活→貯金」用の Excel テンプレートを生成。

設計方針:
- 入力は 3 セルだけ（手取り・先取り投資額・年利）
- 結果は大きな数字とカラーバナーで即可視化
- 5年・10年・20年の複利推移を一目で比較
- iPad 1画面（A4縦想定）で完結

出力: projects/rakuda-sensei/products/digital/先取り投資フロー設計シート.xlsx
      projects/rakuda-sensei/downloads/saki-tori-money-flow-2026.xlsx (添付用)
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT_PRODUCT = ROOT / "products" / "digital" / "先取り投資フロー設計シート.xlsx"
OUT_DOWNLOAD = ROOT / "downloads" / "saki-tori-money-flow-2026.xlsx"

# ===== カラーパレット =====
COLOR_YELLOW = "FFF5C6"   # 入力セル（明るい黄）
COLOR_GREEN = "1F8A4C"    # 達成（深緑）
COLOR_GREEN_BG = "D7F0DD" # 達成バナー背景
COLOR_RED = "C0392B"      # 要見直し（赤）
COLOR_RED_BG = "FADCD9"   # 要見直しバナー背景
COLOR_BLUE_DARK = "1B3A57"  # ヘッダー
COLOR_BLUE_LIGHT = "DCE9F4"  # 結果ハイライト
COLOR_TEAL = "0F5C5F"     # 強調
COLOR_GREY = "F0F0F0"     # 補助


def thin_border():
    s = Side(border_style="thin", color="B0B0B0")
    return Border(left=s, right=s, top=s, bottom=s)


def thick_bottom():
    return Border(bottom=Side(border_style="medium", color=COLOR_BLUE_DARK))


def cell(ws, addr, value, *, font=None, fill=None, align=None, border=None, number_format=None):
    c = ws[addr]
    c.value = value
    if font:
        c.font = font
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if align:
        c.alignment = align
    if border:
        c.border = border
    if number_format:
        c.number_format = number_format
    return c


def merge_and_set(ws, range_str, value, *, font=None, fill=None, align=None, border=None, number_format=None):
    ws.merge_cells(range_str)
    top_left = range_str.split(":")[0]
    cell(ws, top_left, value, font=font, fill=fill, align=align, border=border, number_format=number_format)


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "先取り投資フロー"

    # ===== 列幅 =====
    widths = {"A": 2, "B": 22, "C": 18, "D": 4, "E": 18, "F": 18, "G": 18, "H": 18, "I": 2}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ===== タイトル =====
    merge_and_set(
        ws, "B2:H3",
        "🐪 先取り投資フロー設計シート",
        font=Font(name="Noto Sans CJK JP", size=22, bold=True, color="FFFFFF"),
        fill=COLOR_BLUE_DARK,
        align=Alignment(horizontal="center", vertical="center"),
    )
    merge_and_set(
        ws, "B4:H4",
        "3項目を入れるだけ。あなたの『先取り投資→生活→貯金』の不等式と、5年・10年・20年後の到達額を出します。",
        font=Font(name="Noto Sans CJK JP", size=11, color="555555"),
        align=Alignment(horizontal="center", vertical="center", wrap_text=True),
    )
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 22

    # ===== 入力セクション =====
    merge_and_set(
        ws, "B6:H6",
        "① 黄色セルに3項目入れる",
        font=Font(name="Noto Sans CJK JP", size=13, bold=True, color="FFFFFF"),
        fill=COLOR_TEAL,
        align=Alignment(horizontal="left", vertical="center", indent=1),
    )
    ws.row_dimensions[6].height = 26

    inputs = [
        ("手取り月収 (円)", 250000, "円"),
        ("先取り投資の月額 (円)", 50000, "円"),
        ("想定年利 (%)", 5.0, "％"),
    ]
    for i, (label, default, unit) in enumerate(inputs):
        r = 7 + i
        cell(ws, f"B{r}", label,
             font=Font(name="Noto Sans CJK JP", size=12, bold=True),
             fill=COLOR_GREY,
             align=Alignment(horizontal="left", vertical="center", indent=1),
             border=thin_border())
        cell(ws, f"C{r}", default,
             font=Font(name="Noto Sans CJK JP", size=14, bold=True, color="1B3A57"),
             fill=COLOR_YELLOW,
             align=Alignment(horizontal="right", vertical="center", indent=1),
             border=thin_border(),
             number_format="#,##0" if unit == "円" else "0.0")
        cell(ws, f"D{r}", unit,
             font=Font(name="Noto Sans CJK JP", size=11, color="666666"),
             align=Alignment(horizontal="left", vertical="center"))
        ws.row_dimensions[r].height = 28

    # ===== 自動判定バナー =====
    # 生活費 = 手取り - 先取り投資
    # 不等式: 手取り > 先取り投資 > 0 かつ 残り(生活費) > 0
    r = 11
    merge_and_set(
        ws, f"B{r}:H{r}",
        "② あなたの不等式チェック",
        font=Font(name="Noto Sans CJK JP", size=13, bold=True, color="FFFFFF"),
        fill=COLOR_TEAL,
        align=Alignment(horizontal="left", vertical="center", indent=1),
    )
    ws.row_dimensions[r].height = 26

    # 判定式 (IF):
    # =IF(AND(C7>C8, C8>0, C7-C8>=80000), "緑：不等式成立。このまま自動積立を設定してOK", "赤：先取り投資が手取りに対して大きすぎる or 生活費が足りない。固定費見直しから")
    r = 12
    merge_and_set(
        ws, f"B{r}:H{r+1}",
        '=IF(AND(C7>C8,C8>0,C7-C8>=80000),"✅ 緑バナー：不等式 成立。このまま証券口座で自動積立を設定してOK","🟥 赤バナー：先取り投資が大きすぎ or 生活費 80,000 円を切る。固定費から見直し")',
        font=Font(name="Noto Sans CJK JP", size=15, bold=True, color="FFFFFF"),
        align=Alignment(horizontal="center", vertical="center", wrap_text=True),
    )
    # バナーの背景は条件付き書式で
    from openpyxl.formatting.rule import FormulaRule
    green_fill = PatternFill("solid", fgColor=COLOR_GREEN)
    red_fill = PatternFill("solid", fgColor=COLOR_RED)
    ws.conditional_formatting.add(
        f"B{r}:H{r+1}",
        FormulaRule(formula=[f"AND($C$7>$C$8,$C$8>0,$C$7-$C$8>=80000)"], fill=green_fill),
    )
    ws.conditional_formatting.add(
        f"B{r}:H{r+1}",
        FormulaRule(formula=[f"NOT(AND($C$7>$C$8,$C$8>0,$C$7-$C$8>=80000))"], fill=red_fill),
    )
    ws.row_dimensions[r].height = 28
    ws.row_dimensions[r+1].height = 28

    # ===== お金の流れ可視化 =====
    r = 15
    merge_and_set(
        ws, f"B{r}:H{r}",
        "③ お金の流れ（先取り投資→生活→貯金）",
        font=Font(name="Noto Sans CJK JP", size=13, bold=True, color="FFFFFF"),
        fill=COLOR_TEAL,
        align=Alignment(horizontal="left", vertical="center", indent=1),
    )
    ws.row_dimensions[r].height = 26

    # 3カラム表示: 先取り投資 / 生活費 / 貯金余力
    r = 16
    headers = [
        ("B", "C", "① 先取り投資", "=C8", COLOR_BLUE_LIGHT),
        ("E", "E", "② 生活費（残り）", "=C7-C8", COLOR_BLUE_LIGHT),
        ("G", "H", "③ 貯金 / 米株予備", "=MAX(0,C7-C8-100000)", COLOR_BLUE_LIGHT),
    ]
    for start_col, end_col, label, formula, bg in headers:
        rng = f"{start_col}{r}:{end_col}{r}"
        merge_and_set(ws, rng, label,
                      font=Font(name="Noto Sans CJK JP", size=11, bold=True, color=COLOR_BLUE_DARK),
                      fill=bg,
                      align=Alignment(horizontal="center", vertical="center"),
                      border=thin_border())
        rng2 = f"{start_col}{r+1}:{end_col}{r+1}"
        merge_and_set(ws, rng2, formula,
                      font=Font(name="Noto Sans CJK JP", size=24, bold=True, color=COLOR_BLUE_DARK),
                      fill="FFFFFF",
                      align=Alignment(horizontal="center", vertical="center"),
                      border=thin_border(),
                      number_format='#,##0"円"')
        ws.row_dimensions[r].height = 24
        ws.row_dimensions[r+1].height = 44

    # 矢印（テキスト）
    cell(ws, f"D{r+1}", "→",
         font=Font(name="Noto Sans CJK JP", size=24, bold=True, color=COLOR_BLUE_DARK),
         align=Alignment(horizontal="center", vertical="center"))
    cell(ws, f"F{r+1}", "→",
         font=Font(name="Noto Sans CJK JP", size=24, bold=True, color=COLOR_BLUE_DARK),
         align=Alignment(horizontal="center", vertical="center"))

    # 補足
    r = 19
    merge_and_set(ws, f"B{r}:H{r}",
                  "（生活費の目安は 月 100,000 円。「貯金 / 米株予備」は生活費を引いた残りです）",
                  font=Font(name="Noto Sans CJK JP", size=10, color="888888"),
                  align=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[r].height = 20

    # ===== 複利による未来額 =====
    r = 21
    merge_and_set(
        ws, f"B{r}:H{r}",
        "④ 5年・10年・20年で積み上がる額（年利込み）",
        font=Font(name="Noto Sans CJK JP", size=13, bold=True, color="FFFFFF"),
        fill=COLOR_TEAL,
        align=Alignment(horizontal="left", vertical="center", indent=1),
    )
    ws.row_dimensions[r].height = 26

    # 表ヘッダ
    r = 22
    headers = ["年数", "先取り月額×期間（元本）", "複利込みの到達額", "元本に対する増加"]
    for i, h in enumerate(headers):
        col = get_column_letter(2 + i * 2 if i > 0 else 2)
        # 配置調整: B, D-E, F-G, H
        pass
    # シンプルに列指定
    cell(ws, f"B{r}", "年数",
         font=Font(name="Noto Sans CJK JP", size=11, bold=True, color="FFFFFF"),
         fill=COLOR_BLUE_DARK,
         align=Alignment(horizontal="center", vertical="center"),
         border=thin_border())
    merge_and_set(ws, f"C{r}:D{r}", "元本（積立×期間）",
                  font=Font(name="Noto Sans CJK JP", size=11, bold=True, color="FFFFFF"),
                  fill=COLOR_BLUE_DARK,
                  align=Alignment(horizontal="center", vertical="center"),
                  border=thin_border())
    merge_and_set(ws, f"E{r}:F{r}", "複利込みの到達額",
                  font=Font(name="Noto Sans CJK JP", size=11, bold=True, color="FFFFFF"),
                  fill=COLOR_BLUE_DARK,
                  align=Alignment(horizontal="center", vertical="center"),
                  border=thin_border())
    merge_and_set(ws, f"G{r}:H{r}", "複利で増えた分",
                  font=Font(name="Noto Sans CJK JP", size=11, bold=True, color="FFFFFF"),
                  fill=COLOR_BLUE_DARK,
                  align=Alignment(horizontal="center", vertical="center"),
                  border=thin_border())
    ws.row_dimensions[r].height = 26

    # 5/10/20年の行
    # 複利月次積立公式: FV = PMT * ((1+r/12)^(n*12) - 1) / (r/12)
    # ここで r = C9/100 (年利%), n = 年数, PMT = C8 (月額)
    years_list = [5, 10, 20]
    for i, y in enumerate(years_list):
        r = 23 + i
        cell(ws, f"B{r}", f"{y}年後",
             font=Font(name="Noto Sans CJK JP", size=16, bold=True, color=COLOR_BLUE_DARK),
             fill=COLOR_GREY,
             align=Alignment(horizontal="center", vertical="center"),
             border=thin_border())
        # 元本
        merge_and_set(ws, f"C{r}:D{r}",
                      f"=C8*12*{y}",
                      font=Font(name="Noto Sans CJK JP", size=14, bold=True, color="555555"),
                      align=Alignment(horizontal="center", vertical="center"),
                      border=thin_border(),
                      number_format='#,##0"円"')
        # 複利込み (PMT * ((1+r/12)^(n*12)-1) / (r/12))
        merge_and_set(ws, f"E{r}:F{r}",
                      f"=C8*((1+C9/100/12)^({y}*12)-1)/(C9/100/12)",
                      font=Font(name="Noto Sans CJK JP", size=18, bold=True, color=COLOR_GREEN),
                      align=Alignment(horizontal="center", vertical="center"),
                      border=thin_border(),
                      number_format='#,##0"円"')
        # 複利で増えた分
        merge_and_set(ws, f"G{r}:H{r}",
                      f"=(C8*((1+C9/100/12)^({y}*12)-1)/(C9/100/12)) - C8*12*{y}",
                      font=Font(name="Noto Sans CJK JP", size=14, bold=True, color=COLOR_TEAL),
                      align=Alignment(horizontal="center", vertical="center"),
                      border=thin_border(),
                      number_format='+#,##0"円"')
        ws.row_dimensions[r].height = 38

    # ===== 一言コメント =====
    r = 27
    merge_and_set(
        ws, f"B{r}:H{r}",
        '=CONCATENATE("👉 20年後、元本 ", TEXT(C8*12*20,"#,##0"), "円 が ", TEXT(C8*((1+C9/100/12)^(20*12)-1)/(C9/100/12),"#,##0"), "円 に化けます。複利の威力を体感してください。")',
        font=Font(name="Noto Sans CJK JP", size=12, bold=True, color=COLOR_BLUE_DARK),
        fill=COLOR_BLUE_LIGHT,
        align=Alignment(horizontal="center", vertical="center", wrap_text=True),
        border=thin_border(),
    )
    ws.row_dimensions[r].height = 36

    # ===== 次の一歩 =====
    r = 29
    merge_and_set(
        ws, f"B{r}:H{r}",
        "⑤ 緑バナーが出た方の次の一歩",
        font=Font(name="Noto Sans CJK JP", size=13, bold=True, color="FFFFFF"),
        fill=COLOR_TEAL,
        align=Alignment(horizontal="left", vertical="center", indent=1),
    )
    ws.row_dimensions[r].height = 26

    steps = [
        "1. SBI・楽天・マネックスのいずれかで証券口座を開く（無料）",
        "2. 「つみたて投資枠」で eMAXIS Slim S&P500 を選択",
        "3. 「成長投資枠」で eMAXIS Slim 全世界株式（オルカン）を選択",
        "4. 上の②の金額を 50:50 で割って、毎月の自動積立を設定",
        "5. 設定したら 5 年間ログインしない（手動売買が最大の敵）",
    ]
    for i, txt in enumerate(steps):
        rr = 30 + i
        merge_and_set(ws, f"B{rr}:H{rr}", txt,
                      font=Font(name="Noto Sans CJK JP", size=11, color="333333"),
                      align=Alignment(horizontal="left", vertical="center", indent=2))
        ws.row_dimensions[rr].height = 22

    # ===== 赤バナーが出た方の改善案 =====
    r = 36
    merge_and_set(
        ws, f"B{r}:H{r}",
        "⑥ 赤バナーが出た方の改善案（固定費から見直す）",
        font=Font(name="Noto Sans CJK JP", size=13, bold=True, color="FFFFFF"),
        fill=COLOR_RED,
        align=Alignment(horizontal="left", vertical="center", indent=1),
    )
    ws.row_dimensions[r].height = 26

    fixes = [
        "通信費 → 格安SIM へ（月 8,000 円 → 1,500 円）",
        "保険 → 掛け捨て or 共済へ見直し（月 数千円〜）",
        "サブスク → 90 日使ってないものは即解約",
        "電気・ガス → 新電力プラン比較（年 1〜2 万円）",
        "家賃 → 引っ越し or 家賃交渉（年 数十万円）",
    ]
    for i, txt in enumerate(fixes):
        rr = 37 + i
        merge_and_set(ws, f"B{rr}:H{rr}", txt,
                      font=Font(name="Noto Sans CJK JP", size=11, color="333333"),
                      align=Alignment(horizontal="left", vertical="center", indent=2))
        ws.row_dimensions[rr].height = 22

    # ===== フッター =====
    r = 43
    merge_and_set(ws, f"B{r}:H{r}",
                  "🐪 残業嫌いのらくだ先生　|　note 記事 003 添付　|　数字は1万円単位で四捨五入推奨",
                  font=Font(name="Noto Sans CJK JP", size=10, color="888888"),
                  align=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[r].height = 20

    # ===== シート保護（数式セルを守る・任意） =====
    # 簡単化のため省略

    # ===== ガイドシート（任意） =====
    g = wb.create_sheet("使い方")
    g.column_dimensions["A"].width = 2
    g.column_dimensions["B"].width = 100
    rows = [
        ("🐪 先取り投資フロー設計シート 使い方", 18, True, "FFFFFF", COLOR_BLUE_DARK),
        ("", 11, False, "000000", None),
        ("【3項目を黄色セルに入れるだけ】", 14, True, COLOR_BLUE_DARK, COLOR_BLUE_LIGHT),
        ("① 手取り月収（額面ではなく口座に入る金額）", 11, False, "333333", None),
        ("② 先取り投資の月額（NISA 自動積立に回したい金額）", 11, False, "333333", None),
        ("③ 想定年利（迷ったら 5% のまま。S&P500 の長期平均は 7% ですが円ベース・インフレ調整後で 5%）", 11, False, "333333", None),
        ("", 11, False, "000000", None),
        ("【判定の意味】", 14, True, COLOR_BLUE_DARK, COLOR_BLUE_LIGHT),
        ("✅ 緑：先取り投資後の残りで月8万円以上の生活費が確保できる → そのまま自動積立OK", 11, False, COLOR_GREEN, None),
        ("🟥 赤：生活費が8万円を切る or 先取り投資が手取りより大きい → 固定費見直しから", 11, False, COLOR_RED, None),
        ("", 11, False, "000000", None),
        ("【数値の根拠】", 14, True, COLOR_BLUE_DARK, COLOR_BLUE_LIGHT),
        ("複利計算式：FV = PMT × ((1+r/12)^(n×12) − 1) / (r/12)", 11, False, "333333", None),
        ("（PMT=月額積立、r=年利、n=年数。月複利の積立FV公式）", 11, False, "888888", None),
        ("", 11, False, "000000", None),
        ("【注意】", 14, True, COLOR_BLUE_DARK, COLOR_BLUE_LIGHT),
        ("・実際のリターンは市場により変動します。年利5%は保証ではありません", 11, False, "333333", None),
        ("・本シートは家計設計の参考。投資判断は自己責任でお願いします", 11, False, "333333", None),
        ("・iDeCo・つみたて NISA・小規模企業共済などとの併用は別途検討", 11, False, "333333", None),
    ]
    for i, (txt, size, bold, fg, bg) in enumerate(rows, start=2):
        c = g.cell(row=i, column=2, value=txt)
        c.font = Font(name="Noto Sans CJK JP", size=size, bold=bold, color=fg)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        g.row_dimensions[i].height = max(24, size * 1.6)

    # ===== 保存 =====
    OUT_PRODUCT.parent.mkdir(parents=True, exist_ok=True)
    OUT_DOWNLOAD.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PRODUCT)
    # 同じ内容を downloads/ にもコピー保存
    wb.save(OUT_DOWNLOAD)
    print(f"✅ {OUT_PRODUCT}")
    print(f"✅ {OUT_DOWNLOAD}")


if __name__ == "__main__":
    build()
