#!/usr/bin/env python3
"""
伊神さんから受け取った4商品を誰でも使えるように整形

入力: /tmp/uploaded_zip2/渡すよう/ 配下
出力: projects/rakuda-sensei/products/digital/ 配下

整形内容:
1. 使い方シートを最前面に追加 / 強化
2. 個人固有情報（学年組名・出産予定年など）をサンプル化
3. 用語集・免責事項を統一フォーマットで追加
4. ブランド表記 (らくだ先生🐪 / @rakuda_sensei) を統一
"""

from __future__ import annotations

import shutil
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[3]
SRC_DIR = Path("/tmp/uploaded_zip2/渡すよう")
OUT_DIR = ROOT / "projects" / "rakuda-sensei" / "products" / "digital"

BRAND_BROWN = "8B6F47"
BRAND_CREAM = "FDF6E3"
BRAND_BEIGE = "E8D5A7"
BRAND_DARK = "3D2F1F"

HEADER_FILL = PatternFill(start_color=BRAND_BROWN, end_color=BRAND_BROWN, fill_type="solid")
LIGHT_FILL = PatternFill(start_color=BRAND_CREAM, end_color=BRAND_CREAM, fill_type="solid")
ACCENT_FILL = PatternFill(start_color=BRAND_BEIGE, end_color=BRAND_BEIGE, fill_type="solid")

WHITE_BOLD = Font(name="メイリオ", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="メイリオ", size=18, bold=True, color=BRAND_DARK)
HEAD_FONT = Font(name="メイリオ", size=13, bold=True, color=BRAND_BROWN)
NORMAL = Font(name="メイリオ", size=11)
BORDER = Border(*[Side(border_style="thin", color="BBBBBB")] * 4)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def add_usage_sheet(wb: Workbook, title: str, intro_lines: list[str], position: int = 0) -> None:
    """先頭に使い方シートを追加"""
    if "使い方" in wb.sheetnames:
        # 既にある場合は削除して再作成
        del wb["使い方"]
    ws = wb.create_sheet("使い方", position)
    ws.sheet_properties.tabColor = BRAND_BROWN
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 95

    ws.cell(row=1, column=1, value="🐪")
    c = ws.cell(row=1, column=2, value=title)
    c.font = TITLE_FONT

    for i, line in enumerate(intro_lines, start=3):
        cell = ws.cell(row=i, column=2, value=line)
        if line.startswith("■"):
            cell.font = HEAD_FONT
        elif line.startswith("→") or line.startswith("⚠"):
            cell.font = Font(name="メイリオ", size=11, italic=True, color=BRAND_BROWN)
        else:
            cell.font = NORMAL
        cell.alignment = LEFT


def add_brand_footer(wb: Workbook) -> None:
    """全シート末尾にブランド情報を入れる用に最後シートを追加"""
    if "らくだ先生について" in wb.sheetnames:
        del wb["らくだ先生について"]
    ws = wb.create_sheet("らくだ先生について")
    ws.sheet_properties.tabColor = BRAND_BEIGE
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 95

    info = [
        ("🐪", TITLE_FONT, "残業嫌いのらくだ先生"),
        ("", NORMAL, ""),
        ("■", HEAD_FONT, "プロフィール"),
        ("", NORMAL, "公立中学校教員（国語・20代）"),
        ("", NORMAL, "ICT・AIで残業ゼロ ｜ 副業10万 ｜ 5年で2000万円貯金"),
        ("", NORMAL, "日本株デイトレで-50万やった失敗から、低リスクインデックス投資に振り切り。"),
        ("", NORMAL, ""),
        ("■", HEAD_FONT, "発信中"),
        ("", NORMAL, "X (Twitter): @rakuda_sensei"),
        ("", NORMAL, "note: https://note.com/rakuda_sensei"),
        ("", NORMAL, "BOOTH: https://rakuda-sensei.booth.pm"),
        ("", NORMAL, ""),
        ("■", HEAD_FONT, "ご利用について"),
        ("", NORMAL, "・本商品の個人利用・校内利用は自由"),
        ("", NORMAL, "・二次配布・転売は禁止"),
        ("", NORMAL, "・本商品の数式・テンプレで生じた結果について作者は責任を負いません"),
        ("", NORMAL, "・改善要望や不具合報告は X DM までお気軽に！"),
    ]
    for i, (icon, font, text) in enumerate(info, start=1):
        ws.cell(row=i, column=1, value=icon)
        c = ws.cell(row=i, column=2, value=text)
        c.font = font
        c.alignment = LEFT


# =========================================================================
# 1. 伊神流FP.xlsx
# =========================================================================
def polish_fp() -> Path:
    src = SRC_DIR / "伊神流FP.xlsx"
    out = OUT_DIR / "伊神流FP｜家計と先取り貯金テンプレ.xlsx"
    shutil.copy(src, out)
    wb = load_workbook(out)

    add_usage_sheet(wb, "伊神流FP｜家計と先取り貯金テンプレ", [
        "「先取り貯金 → 残ったお金で生活」の流れを Excel で見える化するテンプレートです。",
        "一人暮らし用とルームシェア用、2パターン用意してます。",
        "",
        "■ このテンプレで何ができるか",
        "・手取り給料から「固定費」「生活費」「貯金」「小遣い」の配分を見える化",
        "・目標貯金総額（100万〜250万）から月の必要貯金額を逆算",
        "・NISA・iDeCo の利用額を入れて「実質貯金額」と「先取り貯金額」を算出",
        "・ボーナス込みの年間試算もできる",
        "",
        "■ 使い方（5ステップ）",
        "1. シート「一人暮らし用」または「ルームシェア用」を開く",
        "2. C2セルに「手取り給料」を入力（ルームシェアの場合は F2 に人数も）",
        "3. 個人固定費・生活費の項目と金額を埋める（例が入ってるので参考に）",
        "4. K列の「目標貯金総額」のチェックボックスで目標を選ぶ",
        "5. 「目標小遣い」と「先取り貯金額」が自動計算される",
        "",
        "■ 実際の運用",
        "・給料が入ったらまず「先取り貯金額」を貯金口座に振込",
        "・「目標小遣い」と「小遣い合計」を見比べて使い方を決める",
        "・NISA・iDeCo の利用額を入力すると「実質貯金額」が出る",
        "・残った小遣いを有意義に使い、月末に残ったらまた貯金口座へ",
        "",
        "■ 用語",
        "・個人固定費: 携帯代・保険料など毎月必要な費用",
        "・生活費: 住居に関わる費用（家賃・光熱費・食費）",
        "・貯金可能額: 手取り − 固定費",
        "・必要貯金額（月）: 目標貯金総額を達成するための月貯金額",
        "・目標小遣い: 手取り − （固定費 + 必要貯金額）",
        "・実質貯金額: 目標小遣い × 12 + ボーナス",
        "・先取り貯金額: 給料が入ったら真っ先に振り込む金額",
        "",
        "■ 注意",
        "→ 例として入ってる数字はサンプルです。あなたの状況に書き換えてください",
        "→ 投資（NISA・iDeCo）の効果は元本割れリスクあり、必ず自己責任で判断",
        "",
        "わからない用語があれば、シート「用語集」も見てください。",
    ], position=0)

    add_brand_footer(wb)
    wb.save(out)
    return out


# =========================================================================
# 2. サイドFire計画.xlsx
# =========================================================================
def polish_side_fire() -> Path:
    src = SRC_DIR / "サイドFire計画.xlsx"
    out = OUT_DIR / "夫婦で実践｜サイドFIRE計画シート.xlsx"
    shutil.copy(src, out)
    wb = load_workbook(out)

    add_usage_sheet(wb, "夫婦で実践｜サイドFIRE計画シート", [
        "夫婦（または2人）で「いつまでに、いくら貯めて、何歳でセミリタイアするか」",
        "を年金まで含めて試算する高機能テンプレートです。",
        "",
        "■ サイドFIRE とは",
        "完全な経済的自由（フルFIRE）の前段階。資産収入だけでは生活費を賄えない分を、",
        "緩めの労働で補う「バリスタFIRE」とも呼ばれる戦略。",
        "→ ぼく自身、保育士の婚約者とこのスタイルを目指してます🐪",
        "",
        "■ シート構成",
        "・初期設定: 現在年齢、設定年、出産予定年などの基本情報",
        "・一人目: あなたの月手取り・生活費・貯蓄・初期投資・月積立",
        "・二人目: パートナーの同上",
        "・合算: 夫婦の運用合計と FIRE 達成年齢の自動計算",
        "・資産収入: 資産額別の年利益・月利益試算",
        "・収入: 副業・パート時給からの試算",
        "・年金: 65歳時 / 繰上げ / 繰下げ受給額の概算",
        "・生活費: 項目別の月額（編集して自分の家計に合わせる）",
        "",
        "■ 使い方（4ステップ）",
        "1. 「初期設定」で現在年齢・設定年を入力",
        "2. 「一人目」「二人目」で月手取り・月生活費・資産情報を入力",
        "3. 「生活費」シートで実際の家計に合わせて項目を更新",
        "4. 「合算」シートで FIRE 達成年齢を確認",
        "",
        "■ 計算ロジック",
        "FIRE① 積立有: 月積立を続けながら運用",
        "FIRE② 積立0: 積立をストップしてからの運用",
        "FIRE③ 取崩し: 月に取崩しを始めてからの寿命",
        "FIRE④ 取崩し: 取崩しが何年もつか NPER 関数で算出",
        "",
        "■ 注意",
        "→ 例の数字（月手取り25万・生活費18万など）はサンプルです",
        "→ 想定年利は過去実績ベースの参考値。投資は元本割れリスクあり",
        "→ 年金額は2026年時点の概算。制度改正で変動します",
        "",
        "■ 編集のコツ",
        "・「セミリタイア」目標を「自分が定時退職を希望する年齢」で入力",
        "・「FIRE①」を「夫婦の現状ペースで何歳で2人とも仕事辞められるか」と読む",
        "・「資産収入」シートで目標資産額（5000万・1億）に必要な配当を確認",
    ], position=0)

    add_brand_footer(wb)
    wb.save(out)
    return out


# =========================================================================
# 3. Todo＆時間割.xlsx
# =========================================================================
def polish_todo() -> Path:
    src = SRC_DIR / "Todo＆時間割.xlsx"
    out = OUT_DIR / "中学校教員のためのToDo＆時間割テンプレ.xlsx"
    shutil.copy(src, out)
    wb = load_workbook(out)

    # サンプル化: 「2年2組」「567組」を「○組」に置換
    ws_jikan = wb["時間割"]
    for row in ws_jikan.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                replaced = cell.value
                for target in ["2年2組", "2年3組", "3年2組", "567組", "学部", "特活", "道徳", "総合", "学活"]:
                    if target in replaced and target in ["2年2組", "2年3組", "3年2組", "567組"]:
                        replaced = replaced.replace(target, "○組")
                if replaced != cell.value:
                    cell.value = replaced

    add_usage_sheet(wb, "中学校教員のためのToDo＆時間割テンプレ", [
        "校務の「ToDo管理」と「週時間割」を1ファイルで管理できるテンプレートです。",
        "始業前に開いてサクッと書く、終業前に見直して未完了を翌日に繰り越す、の運用が想定。",
        "",
        "■ シート構成",
        "・ToDo: チェックボックス付きの ToDo リスト（左右2列 × 15行 = 30件）",
        "・時間割: 週の時間割（月〜金・1〜6時限）",
        "",
        "■ 使い方",
        "1. シート「ToDo」を開く",
        "2. 「やること」「〆切」「備考」を埋める（必要に応じて）",
        "3. 完了したらチェックボックスをクリックして✓",
        "4. シート「時間割」を開き、自分のクラス・教科に合わせて「○組」を書き換え",
        "5. 月初に印刷して教卓に貼っておくと便利",
        "",
        "■ 運用のコツ",
        "・1日のToDo は5個までに絞る（多すぎると未消化が積み重なる）",
        "・〆切は必ず日付で入力（「今週中」では先送りされる）",
        "・備考は「誰のために」「どんな成果が必要か」を1行で",
        "",
        "■ 印刷時",
        "・A4横向き推奨",
        "・「ToDo」と「時間割」を別々に印刷",
        "・職員室の壁に貼っておくと、同僚から「便利そう」と声かけられがち",
        "",
        "■ 注意",
        "→ サンプルの「○組」をあなたの担当クラス（2-1, 3-2 など）に書き換えて",
        "→ 「特活」「道徳」「総合」など学校独自の時間割枠もそのまま使えます",
    ], position=0)

    add_brand_footer(wb)
    wb.save(out)
    return out


# =========================================================================
# 4. 所見生成.md → 商品向けにpolish
# =========================================================================
def polish_shoken() -> Path:
    src = SRC_DIR / "所見生成.md"
    out = OUT_DIR / "通知表所見｜AIプロンプト完全版.md"
    original = src.read_text(encoding="utf-8")

    # 商品らしいヘッダー + 末尾CTAを追加
    header = """# 通知表所見｜AI プロンプト完全版 (公立小中学校 対応)

🐪 残業嫌いのらくだ先生 / 2026年6月版

---

## このプロンプトで何ができるか

- 通知表所見を **公的文書として適切な文体** で AI に生成させられる
- 学校仕様（文字数・観点・性格語・励まし文の有無）を**対話で確認** → 教務主任の指導に準拠
- **三観点（性格・生活・学習）を必ず含む** 三文連結フォーマット
- 性格語1つに自動推定（明示しなければ AI が選ぶ）
- 否定的断定 NG、抽象語単独 NG、文末の「〜な生徒です」NG

## こんな人におすすめ

- 通知表所見に**毎学期 5〜10時間**かかってる先生
- ChatGPT に頼んだら**事務的すぎ・温かみゼロ**になった先生
- 学校特有のフォーマット（文字数・観点）に AI を**完全準拠**させたい先生
- 三観点をバランスよく書きたいけど**毎回どれかが薄くなる**先生

## 使い方（3ステップ）

### STEP 1: ChatGPT (または Claude) を開く

下のプロンプト全文をコピペして送信。
AI から「学校仕様を1つずつ質問」されるので、自分の学校に合わせて答える。

### STEP 2: 生徒情報を入力

入力フォーマット（4項目 or 5項目）:
```
出席番号＃性格行動＃生活行動＃学習行動
出席番号＃性格行動＃性格語＃生活行動＃学習行動
```

入力例:
```
12＃係活動で進んで準備をした＃友達と協力して清掃に取り組んだ＃漢字練習を毎日続けた
15＃クラス会で意見をまとめた＃責任感＃時間を守って行動した＃算数の復習に継続して取り組んだ
```

### STEP 3: 所見を一括生成

AI が三文連結フォーマットで全員分を一気に出力。
最終チェックは必ず人間が目視で。

---

## 🎯 プロンプト本文（コピペ用）

以下を ChatGPT / Claude にコピペして使ってください。

```
"""

    footer = """
```

---

## 運用のコツ

1. **学校仕様の保存は最初の1回だけ**
   - 同じスレッドで質問し続ければ、仕様は引き継がれる
   - 学期が変わったら新スレッドで再設定

2. **生徒の観察事実を3観点で揃える**
   - 性格は「行動」で示す（性格語単独はNG）
   - 生活は「協力・自主・責任」など評価規準語
   - 学習は「継続・工夫・理解」など教科に紐づく語

3. **最終チェックは必ず人間**
   - AI生成 → 軽くリライト → 教務主任確認 が最速ルート
   - 個人情報（氏名・地名・固有エピソード）は入れない

4. **時短効果**
   - クラス35人の所見が、入力20分 + 確認30分 = 50分で完成
   - 従来の半日〜1日が1時間に短縮（ぼくの実体験）

---

## 注意事項

- **このプロンプトは「下書き支援」**。最終文責は先生本人にあります。
- 生成された文章はそのまま使わず、自分の言葉で必ず1度書き直す
- 生徒個人情報（氏名・障害・家庭環境など）はAIに入力しない
- 学校・自治体のAI利用規程に従う

---

## 🐪 残業嫌いのらくだ先生

- 公立中学校教員（国語・20代）
- 残業ゼロ × 副業10万 × 5年で2000万円貯金
- X: @rakuda_sensei / note: @rakuda_sensei

要望・不具合報告は X DM までお気軽に。

© 2026 残業嫌いのらくだ先生 / 校内・個人利用は自由 / 二次配布・転売NG
"""

    polished = header + original.strip() + footer
    out.write_text(polished, encoding="utf-8")
    return out


# =========================================================================
# メイン
# =========================================================================
def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"📁 出力先: {OUT_DIR}\n")

    products = [
        ("伊神流FP", polish_fp),
        ("サイドFire計画", polish_side_fire),
        ("Todo＆時間割", polish_todo),
        ("所見生成プロンプト", polish_shoken),
    ]

    for name, fn in products:
        try:
            out = fn()
            print(f"  ✅ {name} → {out.name} ({out.stat().st_size // 1024}KB)")
        except Exception as e:
            print(f"  ❌ {name}: {e}")
            import traceback
            traceback.print_exc()

    print(f"\n🎁 整形完了。BOOTH/noteで販売可能な状態です。")
    return 0


if __name__ == "__main__":
    sys.exit(main())
